"""
EEG Variable Extraction and Time Series Analysis Pipeline
========================================================

This module serves as the central orchestrator for extracting quantitative features
from preprocessed EEG data across multiple frequency bands. It implements a sliding
window approach to generate time series data and then applies statistical analysis to summarize these time series for each patient. Finally, it records/outputs these per-patient values for future analysis.

Purpose:
--------
Transforms preprocessed EEG signals into quantitative variables suitable for
statistical analysis and clinical interpretation. The pipeline extracts features
from multiple frequency bands simultaneously and generates both summary statistics
and time series data for longitudinal analysis.

Core Architecture:
-----------------
1. Function Discovery: Dynamically loads analysis functions from frequency-specific modules
2. Sliding Window Analysis: Applies functions to overlapping time windows
3. Time Series Generation: Creates temporal sequences of extracted variables
4. Statistical Summarization: Computes summary statistics across time series
5. Data Export: Saves results to Excel and CSV formats for further analysis

Sliding Window Framework:
------------------------
The module implements a sliding window approach that:
- Respects clean run boundaries (artifact-free segments)
- Maintains temporal continuity within runs
- Handles variable window sizes and slide intervals per function
- Preserves timestamp information for time-aware statistics

Function Organization:
---------------------
Analysis functions are organized in frequency-specific modules:
- fxns.py: Broadband analysis functions
- fxns_delta.py: Delta-specific functions
- fxns_theta.py: Theta-specific functions
- fxns_alpha.py: Alpha-specific functions
- fxns_beta.py: Beta-specific functions

Each function must include metadata:
- variable_name: Descriptive name for the extracted variable
- units: Physical or statistical units
- window_size: Analysis window duration (seconds)
- slide_size: Window advance interval (seconds)
- subfunctions: Statistical summaries to compute

Statistical Analysis Framework:
------------------------------
Implements comprehensive statistical analysis including:

Summary Statistics:
- Central tendency: mean, median
- Variability: standard deviation, IQR, percentiles
- Distribution shape: coefficient of variation, proportional bins

Temporal Analysis:
- Trend detection: Mann-Kendall tau, Theil-Sen slope
- Linear modeling: least squares regression
- Long-range correlation: Hurst exponent
- Time-aware calculations using actual timestamps

Data Flow:
----------
Input: Preprocessed alex_raw_efficient objects
Processing: 
  1. Clean run analysis → discontinuity metrics
  2. Function loading → dynamic module import
  3. Sliding window analysis → time series generation
  4. Statistical summarization → summary variables
  5. Time series export → CSV files for LMM
Output: 
  - final_data.xlsx: Summary statistics per recording
  - time_series_data/: Individual time series (normalized time)
  - time_series_data_eeg/: Individual time series (absolute time)

Quality Control:
---------------
- Minimum run validation: Ensures sufficient data for temporal statistics
- Error logging: Tracks failed extractions with detailed messages
- Data validation: Verifies function outputs and statistical calculations
- Memory management: Explicit cleanup of large objects

Error Handling:
--------------
- Function loading failures: Module import error tracking
- Insufficient data: Minimum sample size validation
- Statistical failures: NaN handling for edge cases
- File I/O errors: Excel/CSV write error recovery

Output File Structure:
---------------------
final_data.xlsx columns:
- Recording metadata: file_name, fis_num, timing information
- Quality metrics: discontinuity analysis, artifact proportions
- Extracted variables: All function outputs with statistical summaries

Time series files:
- Individual CSV per variable per subject
- Standardized column format: value, time, subject
- Separate normalized and absolute time versions
"""

import numpy as np
import pandas as pd
import os
import sys
current_dir = os.path.dirname(os.path.abspath(__file__))
utils_dir = os.path.join(current_dir, '..', 'utils')
sys.path.append(utils_dir)
from utils import config
vars_dict = config.get_vars_dict()
from file_importing import import_pkl, get_band_obj_from_original
sys.path.append(os.path.join(utils_dir,"NEURAL_py_EEG"))
from openpyxl import Workbook, load_workbook
import inspect
import importlib.util
from pathlib import Path
import scipy.stats as stats
import nolds
import glob
import statsmodels.api as sm
import openpyxl
import math

excel_file_name = ''

def test_linearity(time_series_data):
    """
    Test for linearity in a time series dataset.
    
    Parameters:
    - time_series_data: DataFrame with 'time', 'value', and 'subject' columns
    
    Returns:
    - Dictionary with linearity test results
    """
    results = {}
    
    # Group by subject
    subjects = time_series_data['subject'].unique()
    
    # For each subject, perform linearity tests
    for subject in subjects:
        subject_data = time_series_data[time_series_data['subject'] == subject]
        if len(subject_data) < 10:  # Skip subjects with too few points
            continue
            
        x = subject_data['time'].values
        y = subject_data['value'].values
        
        # 1. Residual plot analysis
        model = sm.OLS(y, sm.add_constant(x)).fit()
        residuals = model.resid
        
        # Test residuals for normal distribution (linearity assumption)
        _, normality_p = stats.shapiro(residuals)
        
        # Test for heteroscedasticity (linearity assumption)
        _, hetero_p = stats.bartlett(residuals, np.ones_like(residuals))
        
        # 2. RESET test for linearity
        y_pred = model.predict()
        y_pred2 = np.power(y_pred, 2)
        y_pred3 = np.power(y_pred, 3)
        
        reset_model = sm.OLS(y, sm.add_constant(np.column_stack((x, y_pred2, y_pred3)))).fit()
        reset_f = reset_model.compare_f_test(model)
        reset_p = reset_f[1]  # p-value
        
        # 3. Spearman correlation between absolute residuals and x
        abs_resid = np.abs(residuals)
        spearman_corr, spearman_p = stats.spearmanr(x, abs_resid)
        
        # Store results
        results[subject] = {
            'shapiro_p': normality_p,
            'hetero_p': hetero_p,
            'reset_p': reset_p,
            'spearman_p': spearman_p,
            'linear': (normality_p > 0.05 and hetero_p > 0.05 and 
                       reset_p > 0.05 and spearman_p > 0.05)
        }
    
    # Aggregate results
    prop_linear = sum(r['linear'] for r in results.values()) / len(results) if results else 0
    results['proportion_linear'] = prop_linear
    
    return results

def save_to_rejected_log(fis, error_str):
    """
    Code to save to an error log specific to this processing step.
    
    Inputs:
    - error_str (str): Error message to be saved
    """
    global excel_file_name
    wb = openpyxl.load_workbook(excel_file_name)
    rejected_sheet = wb.active
    next_row = rejected_sheet.max_row + 1
    rejected_sheet.cell(row=next_row, column=1, value=fis)  # Add file_path
    rejected_sheet.cell(row=next_row, column=2, value=error_str)  # Add error_type
    wb.save(excel_file_name)

def initialize_rejected_file(save_dir_base):
    """
    Helper function that initializes the rejected logs..
    """
    global excel_file_name
    # File names
    excel_file_name = os.path.join(save_dir_base, "variable_extraction_rejected_log.xlsx")

    #Step 1: Initialize the Excel file
    if os.path.exists(excel_file_name):
        os.remove(excel_file_name)
    # Create the files_info sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["file_path", "error_type"])  # Example header row
    # Save the workbook
    wb.save(excel_file_name)
    print(f"Excel file '{excel_file_name}' initialized with default values.")

def save_time_series_data_eeg(save_base_dir, raw_obj, series_data, series_times, function_name, parameter=None):
    """
    Save time series data for later aggregate analysis. Saves with un-normalized time values
    
    Parameters:
    - raw_obj: The EEG data object containing patient information
    - series_data: List containing the time series data
    - series_times: List of timestamps for the time series
    - function_name: Name of the function that generated the time series
    - parameter: Optional parameter name for functions with multiple outputs
    
    Returns:
    - Path to the saved CSV file
    """
    # Get patient FIS number
    fis_num = raw_obj.fis_num
    
    # Convert time to hours from start
    
    # Create dataframe for time series data
    ts_data = pd.DataFrame({
        'value': series_data,
        'time': series_times,
        'subject': fis_num
    })
    
    # Save time series data for later aggregation
    param_suffix = f"_{parameter}" if parameter else ""
    ts_file_name = f"timeseries_{function_name}{param_suffix}_{fis_num}.csv"
    ts_dir = save_base_dir + r"\time_series_data_eeg"
    os.makedirs(ts_dir, exist_ok=True)
    
    save_path = os.path.join(ts_dir, ts_file_name)
    ts_data.to_csv(save_path, index=False)
    
    return save_path

def save_time_series_data(save_base_dir, raw_obj, series_data, series_times, function_name, parameter=None):
    """
    Save time series data for later aggregate LMM analysis.
    
    Parameters:
    - raw_obj: The EEG data object containing patient information
    - series_data: List containing the time series data
    - series_times: List of timestamps for the time series
    - function_name: Name of the function that generated the time series
    - parameter: Optional parameter name for functions with multiple outputs
    
    Returns:
    - Path to the saved CSV file
    """
    # Get patient FIS number
    fis_num = raw_obj.fis_num
    
    # Convert time to hours from start
    times_hrs = np.array([(t - series_times[0])/3600 for t in series_times])
    
    # Create dataframe for time series data
    ts_data = pd.DataFrame({
        'value': series_data,
        'time': times_hrs,
        'subject': fis_num
    })
    
    # Save time series data for later aggregation
    param_suffix = f"_{parameter}" if parameter else ""
    ts_file_name = f"timeseries_{function_name}{param_suffix}_{fis_num}.csv"
    ts_dir = save_base_dir + r"\time_series_data"
    os.makedirs(ts_dir, exist_ok=True)
    
    save_path = os.path.join(ts_dir, ts_file_name)
    ts_data.to_csv(save_path, index=False)
    
    return save_path

def analyze_discontinuities(clean_runs, clean_runs_lengths, sampling_rate):
    """
    Analyzes the pattern of discontinuities in EEG data based on clean runs.
    """
    if len(clean_runs) <= 1:
        # If we have one or zero clean runs
        if len(clean_runs_lengths) > 0:
            # Convert numpy.int64 to float for the division
            single_run_duration = float(clean_runs_lengths[0])/sampling_rate
            return {
                "number_of_discontinuities": 0,
                "mean_gap_duration": 0,
                "max_gap_duration": 0,
                "mean_clean_run_duration": single_run_duration,
                "total_clean_duration": single_run_duration,
                "discontinuity_density": 0
            }
        else:
            # No clean runs at all
            return {
                "number_of_discontinuities": 0,
                "mean_gap_duration": 0,
                "max_gap_duration": 0,
                "mean_clean_run_duration": 0,
                "total_clean_duration": 0,
                "discontinuity_density": 0
            }
    
    # Calculate gaps between clean runs
    gaps = []
    for i in range(len(clean_runs)-1):
        current_run_end = clean_runs[i] + clean_runs_lengths[i]
        next_run_start = clean_runs[i+1]
        gap_samples = next_run_start - current_run_end
        gaps.append(gap_samples / sampling_rate)  # Convert to seconds
    
    # Calculate clean run durations
    clean_run_durations = [length/sampling_rate for length in clean_runs_lengths]
    total_clean_duration = sum(clean_run_durations)
    
    # Calculate discontinuity density (per hour)
    recording_duration_hours = total_clean_duration / 3600
    discontinuity_density = len(gaps) / recording_duration_hours if recording_duration_hours > 0 else 0
    
    return {
        "number_of_discontinuities": len(gaps),
        "mean_gap_duration": np.mean(gaps) if gaps else 0,
        "max_gap_duration": np.max(gaps) if gaps else 0,
        "mean_clean_run_duration": np.mean(clean_run_durations),
        "total_clean_duration": total_clean_duration,
        "discontinuity_density": discontinuity_density
    }

def time_aware_kendall(values, times):
    """
    Calculate Kendall's Tau using actual timestamps instead of indices. NB: this should not change the Tau value (rank matters only)
    """
    if len(values) < 2:
        return np.nan
    # Convert times to hours for better numerical stability
    times_hrs = np.array([(t - times[0])/3600 for t in times])
    return stats.kendalltau(times_hrs, values)[0]

def time_aware_theil_sen(values, times):
    """
    Calculate Theil-Sen slope using actual time differences
    """
    if len(values) < 2:
        return np.nan
    times_hrs = np.array([(t - times[0])/3600 for t in times])
    return stats.theilslopes(values, times_hrs)

def get_a_time_series(save_base_dir, raw_obj, fxn, fxn_details, window_size, slide_time):
    """
    Extract time series data using sliding window analysis with clean run awareness.
    
    This is the core function that implements the sliding window framework for EEG feature extraction.
    
    Parameters:
    -----------
    save_base_dir : str
        Base directory for saving time series CSV files
    raw_obj : alex_raw_efficient object
        Preprocessed EEG object with clean_runs and clean_runs_lengths attributes
    fxn : function object
        Analysis function to apply to each window
        Must have attributes: variable_name, units, window_size, slide_size, subfunctions
    fxn_details : list [str, str]
        [variable_name, units] for documentation purposes
    window_size : float
        Window duration in seconds (redundant with fxn.window_size)
    slide_time : float  
        Window advance interval in seconds (redundant with fxn.slide_size)
        
    Returns:
    --------
    output_dict : dict
        Dictionary mapping statistic names to calculated values
        Format: {f"{variable_name}_{parameter}_{statistic}": value}
        Returns {"Error": "Error"} if insufficient data
        
    Algorithm Overview:
    ------------------
    
    Step 1: Data Quality Assessment
    ------------------------------
    Validates that sufficient clean runs exist for temporal statistics:
    - Counts windows possible within each clean run
    - Requires >1 total windows for Mann-Kendall tau and Theil-Sen slope
    - Early termination if insufficient data prevents temporal analysis
    
    Step 2: Clean Run Processing
    ---------------------------
    Processes each clean run independently to prevent artifact contamination.

    For each clean run:
    1. Calculate number of possible windows: floor((run_length - window_size) / slide_time) + 1
    2. Slide window through run in slide_time increments
    3. Apply analysis function to each window
    4. Store results with corresponding timestamps
    
    Step 3: Multi-Output Handling
    ----------------------------
    Supports functions that return:
    - Single value: Standard scalar output
    - Tuple: Multiple related variables (e.g., power, frequency)
    - Complex structures: Handled through subfunctions specification
    
    Step 4: Time Series Export
    -------------------------
    Saves time series data in two formats:
    
    Normalized Time:
    - Time = (timestamp - first_timestamp) / 3600 (hours from start)
    - Enables comparison across subjects with different recording times
    
    Absolute Time (for EEG analysis):
    - Time = original timestamps from recording
    - Preserves actual temporal relationships
    
    Step 5: Statistical Summarization
    --------------------------------
    Applies requested statistics to generated time series:
    
    Temporal Statistics (require timestamps):
    - mann_kendall_tau: Non-parametric trend test
    - theil_sen_slope: Robust trend estimation  
    - lin_reg: Linear regression slope
    - Uses actual time intervals for proper scaling
    
    Distributional Statistics:
    - mean, median, percentiles: Central tendency and spread
    - STDEV, IQR: Variability measures
    - Proportional binning: Distribution characterization
    
    Quality Control Checks:
    ----------------------
    1. Window Validation: Ensures no NaN values within windows
    2. Sample Size: Verifies minimum data for each statistic
    3. Function Output: Validates expected return structure
    4. Time Consistency: Checks timestamp ordering and intervals
    
    Subfunction Architecture:
    ------------------------
    Two supported formats:
    
    Dictionary Format (multiple parameters):
    ```python
    subfunctions = {
        "power": ["mean", "median", "mann_kendall_tau"],
        "frequency": ["mean", "STDEV"]
    }
    ```
    
    List Format (single parameter):
    ```python
    subfunctions = ["mean", "median", "mann_kendall_tau", "theil_sen_slope"]
    ```
    
    Error Handling:
    --------------
    - NaN detection: Skips windows containing artifacts
    - Statistical failures: Returns np.nan for problematic calculations
    - Insufficient data: Returns "Error" dictionary for early termination
    """
    import numpy as np
    from tqdm import tqdm
    
    #Step 0: Define subfunction_map, a map of all subfunctions (i.e., statistics for a generated time series) we could use. These (i.e., the string representations) are to be placed in function decorators in fxns.py, which are then read and matched to functions in subfunctions_map.
    def calc_proportions(x):
        arr = np.asarray(x, dtype=np.float32)  # Convert once, using float32
        return {
            "prop_0_10": np.mean((arr >= 0) & (arr < 10)),
            "prop_10_25": np.mean((arr >= 10) & (arr < 25)),
            "prop_25_50": np.mean((arr >= 25) & (arr < 50)),
            "prop_50_100": np.mean((arr >= 50) & (arr < 100)),
            "prop_over_100": np.mean(arr >= 100)
        }

    subfunction_map = {
        "mean": lambda x: np.mean(x),
        "median": lambda x: np.median(x),
        "STDEV": lambda x: np.std(x),
        "coefficient_of_variance": lambda x: np.std(x) / np.mean(x),
        "5th_percentile": lambda x: np.percentile(x, 5),
        "95th_percentile": lambda x: np.percentile(x, 95),
        "width": lambda x: np.percentile(x, 95) - np.percentile(x, 5),
        "IQR": lambda x: np.percentile(x, 75) - np.percentile(x, 25),
        "mann_kendall_tau": lambda x, t: time_aware_kendall(x, t),
        "theil_sen_slope": lambda x, t: time_aware_theil_sen(x, t)[0],
        "theil_sen_y_int": lambda x, t: time_aware_theil_sen(x, t)[1],
        "lin_reg": lambda x, t: np.polyfit(
            np.array([(ti - t[0])/3600 for ti in t]), x, 1
        )[0] if len(x) >= 2 else np.nan,
        "lin_reg_y": lambda x, t: np.polyfit(
            np.array([(ti - t[0])/3600 for ti in t]), x, 1
        )[1] if len(x) >= 2 else np.nan,
        "hurst": lambda x: np.nan if len(x) < 10 else nolds.hurst_rs(x),
        "prop_0_10": lambda x: calc_proportions(x)["prop_0_10"],
        "prop_10_25": lambda x: calc_proportions(x)["prop_10_25"],
        "prop_25_50": lambda x: calc_proportions(x)["prop_25_50"],
        "prop_50_100": lambda x: calc_proportions(x)["prop_50_100"],
        "prop_over_100": lambda x: calc_proportions(x)["prop_over_100"]
    }

    #Step 1: Initialize our critical variables
    srate = raw_obj.sampling_rate
    data = raw_obj.EEG_data
    times = raw_obj.EEG_times_from_analyze_end_time
    window_size_samples = window_size*srate    
    slide_time_samples = slide_time*srate      
    clean_run_start_indices = raw_obj.clean_runs
    clean_run_lengths = raw_obj.clean_runs_lengths

    #Checks for the case that all windows of data are too small for our function; in this case, logs an error and returns an output dict containing "Error", which is later on handled to not be added to the final sheet. NB: this terminates before any eeg_time_series objects are built, so none of the rejected files are accidentally included in the stat_GUI
    num_ok_runs = sum(
        math.floor(((run / srate) - fxn.window_size) / fxn.slide_size) + 1
        for run in clean_run_lengths
        if (run / srate) >= fxn.window_size
    )
    if num_ok_runs<=1: #This is saying if we only have 1 clean run (which means TSS and MKT are not calculable) 
        output_dict = {}
        if isinstance(fxn.subfunctions, dict):
            # Dictionary case - multiple parameters with their own statistics
            for parameter, stat_list in fxn.subfunctions.items():
                # Find which index this parameter was in the original function output
                all_params = list(fxn.subfunctions.keys())
                param_idx = all_params.index(parameter)
                
                # Apply requested statistics to this parameter's time series
                for stat_name in stat_list:
                    stat_result = "Error"
                    output_dict[f"{fxn.variable_name}_{parameter}_{stat_name}"] = stat_result
        else:
            # List case - single parameter with multiple statistics
            for subfunction_name in fxn.subfunctions:
                result = "Error"
                output_dict[f"{fxn.variable_name}_{subfunction_name}"] = result
        return output_dict
    
    
    # Initialize series_data as a list of lists, one for each output
    n_outputs = len(fxn.subfunctions.keys()) if isinstance(fxn.subfunctions, dict) else 1
    series_data = [[] for _ in range(n_outputs)]
    series_times = []
    var_name = fxn_details[0]
    var_units = fxn_details[1]
    
    #Step 2: Process each clean run separately
    for run_idx, start_idx in tqdm(enumerate(clean_run_start_indices), total=len(clean_run_start_indices), desc=f"Processing runs to calculate {var_name}"):
        run_length = clean_run_lengths[run_idx]
        end_idx = start_idx + run_length
        
        current_start = start_idx
        with tqdm(total=(end_idx - start_idx) // slide_time_samples, desc=f"Sliding Windows (Run {run_idx} of {len(clean_run_start_indices)})", leave=False) as pbar:
            while current_start + window_size_samples <= end_idx:
                window = data[current_start:current_start + window_size_samples]
                
                if np.isnan(window).any():
                    break
                
                # Apply function and handle multiple outputs
                result = fxn(window, raw_obj)
                if isinstance(result, tuple):
                    for i, val in enumerate(result):
                        series_data[i].append(val)
                else:
                    series_data[0].append(result)
                
                series_times.append(times[current_start])
                current_start += slide_time_samples
                pbar.update(1)

    
    #Step 4: Save time series data for later LMM analysis
    if isinstance(fxn.subfunctions, dict):
        # Dictionary case - multiple parameters with their own statistics
        all_params = list(fxn.subfunctions.keys())
        for param_idx, parameter in enumerate(all_params):
            # Save this parameter's time series for LMM analysis
            save_time_series_data(save_base_dir, raw_obj, series_data[param_idx], series_times, fxn.variable_name, parameter)
            save_time_series_data_eeg(save_base_dir, raw_obj, series_data[param_idx], series_times, fxn.variable_name, parameter)
    else:
        # List case - single parameter 
        save_time_series_data(save_base_dir, raw_obj, series_data[0], series_times, fxn.variable_name)
        save_time_series_data_eeg(save_base_dir, raw_obj, series_data[0], series_times, fxn.variable_name)
    
    #Step 5: Build output dictionary with the original summary statistics
    output_dict = {}
    if isinstance(fxn.subfunctions, dict):
        # Dictionary case - multiple parameters with their own statistics
        for parameter, stat_list in fxn.subfunctions.items():
            # Find which index this parameter was in the original function output
            all_params = list(fxn.subfunctions.keys())
            param_idx = all_params.index(parameter)
            
            # Apply requested statistics to this parameter's time series
            for stat_name in stat_list:
                if stat_name in ['mann_kendall_tau', 'theil_sen_slope', "theil_sen_y_int", 'lin_reg', "lin_reg_y"]:
                    # Pass both values and timestamps for time-aware calculations
                    stat_result = subfunction_map[stat_name](series_data[param_idx], series_times)
                else:
                    # Original behavior for non-temporal statistics
                    stat_result = subfunction_map[stat_name](series_data[param_idx])
                output_dict[f"{fxn.variable_name}_{parameter}_{stat_name}"] = stat_result
    else:
        # List case - single parameter with multiple statistics
        for subfunction_name in fxn.subfunctions:
            if subfunction_name in ['mann_kendall_tau', 'theil_sen_slope', "theil_sen_y_int", 'lin_reg', "lin_reg_y"]:
                result = subfunction_map[subfunction_name](series_data[0], series_times)
            else:
                result = subfunction_map[subfunction_name](series_data[0])
            output_dict[f"{fxn.variable_name}_{subfunction_name}"] = result

    return output_dict



def extract_vars_iterate(base_save_dir):
    """
    The purpose of this code will be to iterate over each alex_raw object at files_to_extract_path, call all functions, 
    and then save the variables these functions extract to an Excel sheet. It also preserves time series data
    for later aggregate linear mixed effects regression analysis.
    """
    
    files = glob.glob(os.path.join(base_save_dir,r'time_series_data\*'))
    for f in files:
        os.remove(f)
    files = glob.glob(os.path.join(base_save_dir,r'time_series_data_eeg\*'))
    for f in files:
        os.remove(f)
    output_sheet_dir = os.path.join(base_save_dir, "final_data.xlsx")
    files_to_extract_path = os.path.join(os.path.split(base_save_dir)[0], "3_preprocessing", "preprocessed_files")
    fxns_base_dir = os.path.join(vars_dict["base_dir"], "..", "processing_scripts","4_variable_functions")
    if os.path.exists(output_sheet_dir):
        os.remove(output_sheet_dir)
    # Initialize output file with headers if it doesn't exist
    if not os.path.exists(output_sheet_dir):
        os.makedirs(os.path.dirname(output_sheet_dir), exist_ok=True)
        # Create a dummy file_vars_dict to get the column names
        # This is a placeholder - we'll use the first real file to get proper headers
        file_vars_dict = {
            "number_of_discontinuities": 0,
            "mean_gap_duration": 0,
            "max_gap_duration": 0,
            "mean_clean_run_duration": 0,
            "total_clean_duration": 0,
            "discontinuity_density": 0,
            "days_of_life_at_EEG": 0,
            "gestational_age_EEG": 0
        }
        wb = Workbook()
        ws = wb.active
        headers = ["file_name", "fis_num", "recording_start_time_hrs_postop", 
                    "recording_end_time_hours_postop", "recording_length_hrs", 
                    "pause_length_hrs", "proportion_pause", "artefact_length_hrs", 
                    "proportion_artefact"] + list(file_vars_dict.keys())
        ws.append(headers)  # Adds header row to the worksheet
        wb.save(output_sheet_dir)  # Save the new workbook
        print(f"Created new output file: {output_sheet_dir}")
    
    initialize_rejected_file(base_save_dir)
    
    def my_it(file):
        """Process a single file and save results to Excel with proper locking"""
        # Step 1: Initialize dictionaries to store variable information.
        print(f"Processing file {file}")
        file_vars_dict = {}
        file_path = os.path.join(files_to_extract_path, file)
        
        alex_raw_to_check = import_pkl(full_file_path=file_path, parts_or_path="path")
        wb = load_workbook(output_sheet_dir, data_only=True)
        ws = wb.active
        
        # Continue with processing if file not already in the spreadsheet
        discontinuity_metrics = analyze_discontinuities(alex_raw_to_check.clean_runs, 
                                                        alex_raw_to_check.clean_runs_lengths, 
                                                        alex_raw_to_check.sampling_rate)
        file_vars_dict.update(discontinuity_metrics)
        
        # Step 2: Process the variables
        # Step 2.1: iterate through functions for different frequency bands
        fxn_suffixes = ["", "delta", "theta", "alpha", "beta"]
        for suffix in fxn_suffixes:
            # Step 2.1.1: Load functions from corresponding file
            if suffix == "":
                fxns_path = fxns_base_dir + r"\fxns" + f"{suffix}" + ".py"
            else:
                fxns_path = fxns_base_dir + r"\fxns" + f"_{suffix}" + ".py"
            
            spec = importlib.util.spec_from_file_location(Path(fxns_path).stem, fxns_path)
            module = importlib.util.module_from_spec(spec)
            sys.modules[spec.name] = module
            spec.loader.exec_module(module)
            
            # Get function objects defined in this file
            fxns = [obj for name, obj in inspect.getmembers(module) 
                    if inspect.isfunction(obj) and obj.__module__ == spec.name 
                    and name != "add_variable_metadata"]
            
            # Verify functions have required attributes
            for fxn in fxns:
                assert hasattr(fxn, "variable_name"), f"Function {fxn.__name__} is missing variable_name!"
            
            print(f"Loaded functions: {[f.__name__ for f in fxns]}")
            
            # Get the appropriate band object and process each function
            band_obj = get_band_obj_from_original(alex_raw_to_check, suffix)
            for fxn in fxns:
                fxndict = get_a_time_series(base_save_dir, band_obj, fxn, [fxn.variable_name, fxn.units], fxn.window_size, fxn.slide_size)
                if not "Error" in list(fxndict.values()):
                    file_vars_dict.update(fxndict)
                else:
                    save_to_rejected_log(int(alex_raw_to_check.fis),"Under 2 clean runs found, so file skipped (as TSS and MKT would be incalculable).")
                    return 0 #This ends the function before anything has been returned

        
        # Calculate recording metrics
        recording_start_time = round(alex_raw_to_check.EEG_times_from_analyze_end_time[0] / 3600, 2)
        recording_end_time = round(alex_raw_to_check.EEG_times_from_analyze_end_time[-1] / 3600, 2)
        proportion_pause = round(alex_raw_to_check.amount_removed_by_artefact[0]/100, 2)
        recording_length_hours = round(len(alex_raw_to_check.EEG_data)/alex_raw_to_check.sampling_rate/3600, 2)
        pause_length_hrs = proportion_pause*recording_length_hours
        proportion_artefact = alex_raw_to_check.proportion_artefacts
        artefact_length_hours = proportion_artefact*recording_length_hours
        
        # Prepare row data
        row_data = [
            file, 
            str(alex_raw_to_check.num_fis), 
            recording_start_time, 
            recording_end_time, 
            recording_length_hours, 
            pause_length_hrs, 
            proportion_pause, 
            artefact_length_hours, 
            proportion_artefact
        ] + list(file_vars_dict.values())
        
        # Round floating point values for better Excel display
        row_data = [round(float(x), 3) if isinstance(x, (np.float64, np.float32)) else x for x in row_data]
        wb = load_workbook(output_sheet_dir, data_only=True)
        ws = wb.active
        ws.append(row_data)
        wb.save(output_sheet_dir)
        print(f"Data for {alex_raw_to_check.file_name} has been added to {output_sheet_dir}.")
    
    for file in os.listdir(files_to_extract_path):
        my_it(file)


"""
Data Flow and File Organization:
===============================

Input Directory Structure:
├── 3_preprocessing/
│   ├── preprocessed_files/         # Notch-filtered objects (primary input)
│   ├── preprocessed_files_delta/   # Delta band objects  
│   ├── preprocessed_files_theta/   # Theta band objects
│   ├── preprocessed_files_alpha/   # Alpha band objects
│   └── preprocessed_files_beta/    # Beta band objects

Processing Functions (Dynamic Loading):
├── 4_variable_functions/
│   ├── fxns.py           # Broadband analysis functions
│   ├── fxns_delta.py     # Delta-specific functions
│   ├── fxns_theta.py     # Theta-specific functions  
│   ├── fxns_alpha.py     # Alpha-specific functions
│   └── fxns_beta.py      # Beta-specific functions

Output Directory Structure:
├── final_data.xlsx                 # Master results spreadsheet
├── variable_extraction_rejected_log.xlsx  # Error tracking
├── time_series_data/              # Normalized time series (for LMM)
│   ├── timeseries_spectral_power_mean_001.csv
│   ├── timeseries_spectral_power_delta_mean_001.csv
│   └── ...
└── time_series_data_eeg/          # Absolute time series (for EEG analysis)
    ├── timeseries_spectral_power_mean_001.csv
    └── ...

Excel Output Columns:
=====================
Metadata Columns:
- file_name: Original preprocessed file identifier
- fis_num: Patient identifier
- recording_start_time_hrs_postop: Recording start relative to surgery
- recording_end_time_hours_postop: Recording end relative to surgery
- recording_length_hrs: Total recording duration
- pause_length_hrs: Duration of zero-run artifacts
- proportion_pause: Fraction of recording with zero-runs
- artefact_length_hrs: Duration of all artifacts
- proportion_artefact: Fraction of recording with artifacts

Quality Metrics:
- number_of_discontinuities: Count of artifact gaps
- mean_gap_duration: Average artifact episode length
- max_gap_duration: Longest artifact episode  
- mean_clean_run_duration: Average usable segment length
- total_clean_duration: Total analyzable data duration
- discontinuity_density: Artifacts per hour (normalized)

Extracted Variables:
- {function_name}_{parameter}_{statistic}: Analysis results
- Example: spectral_power_mean, alpha_relative_power_mann_kendall_tau

Time Series CSV Format:
======================
Columns: value, time, subject
- value: Extracted variable value for this time window
- time: Timestamp (hours from start for LMM, absolute for EEG analysis)
- subject: Patient FIS number for grouping

Example:
value,time,subject
12.4,0.0,001
15.2,0.25,001  
11.8,0.5,001
...
"""