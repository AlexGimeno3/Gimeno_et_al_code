"""
OBM (Olympic Brainz Monitor) EEG Data Processing Pipeline

This module processes EEG recordings from OBM devices, stitching together multiple .edf files
that are exported in EDF+C format using EDFbrowser. Unlike Nicolet recordings, OBM files use
a different naming convention and pause detection methodology based on measurement timestamps.

Purpose:
--------
Processes OBM EEG recordings by combining segmented .edf files into unified alex_raw_efficient
objects for analysis. OBM recordings are pre-processed through EDFbrowser to create EDF+C format
files with standardized naming (XXX_YYYY.edf format where YYYY is a 4-digit sequence number).

Key Differences from Nicolet Pipeline:
-------------------------------------
1. File Structure: Uses channel-specific subdirectories (CrossEeg, LeftEeg, RightEeg)
2. Naming Convention: Files follow XXX_YYYY.edf pattern instead of simple numeric suffixes
3. Pause Detection: Uses measurement timestamps (meas_date) rather than annotation parsing
4. Channel Selection: Single channel processing per run (CrossEeg/LeftEeg/RightEeg)

Data Structures:
---------------
1. alex_raw: Core EEG data object with voltage arrays and metadata
2. alex_raw_efficient: Memory-optimized version for storage
3. voltages_arr: Accumulated voltage data from all segments
4. meas_start_datetime_arr: Array of measurement start times for pause calculation
5. meas_end_datetime_arr: Array of measurement end times for gap detection
6. pause_indices: Indices where recording pauses occurred

Processing Workflow:
-------------------
1. Navigate to channel-specific subdirectory (e.g., FIS123/CrossEeg/)
2. Sort .edf files by 4-digit suffix (custom_sort handles XXX_YYYY format)
3. For each file, extract meas_date timestamp from raw_EDF.info
4. Calculate pauses as time gaps between consecutive recordings
5. Insert zero-filled arrays during pause periods
6. Concatenate voltage data into continuous stream
7. Build timestamp arrays and extract analysis windows
8. Save as alex_raw_efficient object

Pause Detection Logic:
---------------------
Unlike Nicolet (which parses annotations), OBM uses measurement timestamps:
- meas_start_datetime: Start time from EDF header (raw_EDF.info["meas_date"])
- meas_end_datetime: Calculated as start + signal_length
- pause_length: Time gap between end of previous file and start of current file
- Zero-padding: Gaps filled with zeros for artifact rejection pipeline

File Structure Expected:
-----------------------
OBM_folder/
├── FIS123a/
│   ├── CrossEeg/
│   │   ├── recording_0001.edf
│   │   ├── recording_0002.edf
│   │   └── ...
│   ├── LeftEeg/
│   └── RightEeg/
└── FIS124b/
    └── ...

Configuration Dependencies:
--------------------------
- OBM_channel: Channel selection (["CrossEeg"], ["LeftEeg"], or ["RightEeg"])
- OBM_folder: Base directory containing FIS folders
- OBM_recording_time_data_path: Excel metadata file path
- windows: Analysis time windows (e.g., [[20,24]] for 20-24h post-surgery)
- downsample_rate: EEG downsampling factor

Error Handling:
--------------
- Missing channel directories: Logged and skipped
- Corrupted .edf files: Comprehensive error tracking
- Invalid measurement timestamps: DateTime validation and error logging
- Missing surgery times: Graceful failure with error recording

Output Files:
------------
- stitched_cut_efficient/: Processed alex_raw_efficient objects
- OBM_raw_file_timestamps.xlsx: Processing metadata and timing
- OBM_error_log.xlsx/.txt: Error tracking and debugging information
"""

import os
import sys
current_dir = os.path.dirname(os.path.abspath(__file__))
utils_dir = os.path.join(current_dir, '..', 'utils')
sys.path.append(utils_dir)
from WindowBuildError import WindowBuildError
from file_importing import import_edf_from_path
from alex_raw import alex_raw
from alex_raw_efficient import alex_raw_efficient
import pandas as pd
import re
import openpyxl
from datetime import datetime, timedelta
import numpy as np
from copy import deepcopy
from utils import config
vars_dict = config.get_vars_dict()

# Modify record_file_info to use the lock
def record_file_info(file_path, EEG_times_arr, fis, save_dir_base, chunk = "-", stitched_path=None):
    """
    Records file information and timestamps in an Excel file if the file hasn't been recorded yet.
    
    Args:
        file_path (str): Path to the file being processed
        EEG_times_from_analyze_end_times (array): Array of timestamps
        fis (str): FIS identifier
    """
    excel_path = save_dir_base + r"\OBM_raw_file_timestamps.xlsx"
    # Create Excel file if it doesn't already exist
    if not os.path.exists(excel_path):
        df = pd.DataFrame(columns=['file_name', 'fis', 'start_time', 'end_time'])
        df.to_excel(excel_path, index=False)
    # Read existing data
    df = pd.read_excel(excel_path)    
    # Get start and end times
    start_time = EEG_times_arr[0]
    end_time = EEG_times_arr[1]
    # Create new row
    new_row = {
        'file_name': file_path,
        'fis': fis,
        'start_time': start_time,
        'end_time': end_time,
        'chunk' : chunk,
        'stitched_path' : stitched_path
    }
    # Append new row
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    # Save updated DataFrame
    df.to_excel(excel_path, index=False)
    print(f"Added {file_path} (FIS{fis}) to {excel_path}")

def initialize_error_files(save_dir_base, clear=True):
    """
    Helper function that initializes the Excel file error_log_OBM.xlsx and the text file error_log_OBM.txt.
    """
    # File names
    excel_file_name = save_dir_base + r"\OBM_error_log.xlsx"
    text_file_name = save_dir_base + r"\OBM_error_log.txt"
    today_date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    #Step 1: Initialize the Excel file
    # Create the files_info sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["file_path", "error_type"])  # Example header row
    # Save the workbook
    wb.save(excel_file_name)
    print(f"Excel file '{excel_file_name}' initialized with default values.")

    #Step 2: Initialize or clear the text file
    if clear or not os.path.exists(text_file_name):
        with open(text_file_name, "w") as log_file:
            log_file.write(f"{today_date_time}: error log initialized.\n")
        print(f"Text file '{text_file_name}' initialized with the current date and time.")
    else:
        print(f"Text file '{text_file_name}' already exists. No changes made.")
    
def record_error(error_string, save_dir_base):
    """
    Helper function that builds an error log for this fxn.

    Inputs:
    - error_string (str): the string we want added to the error log
    """
    # Define the log file path
    log_file_path = save_dir_base + r"\OBM_error_log.txt"
    with open(log_file_path, "a") as log_file:
        # Write log initialization with today's date
        log_file.write(f"{error_string}\n\n")

def add_error_to_excel(error_type, file_path, save_dir_base):
    """
    Helper function that updates the information in error_log.xlsx any time there is an error.
    """
    file_name = save_dir_base + r"\OBM_error_log.xlsx"
    
    wb = openpyxl.load_workbook(file_name)
    files_info_sheet = wb.active
    next_row = files_info_sheet.max_row + 1
    files_info_sheet.cell(row=next_row, column=1, value=file_path)  # Add file_path
    files_info_sheet.cell(row=next_row, column=2, value=error_type)  # Add error_type
    wb.save(file_name)
    print(f"Error of type '{error_type}' logged for file: {file_path}")

def clean_up_logs_and_excel(save_dir_base):
    # File paths
    error_log_txt = save_dir_base + r"\OBM_error_log.txt"
    error_log_xlsx = save_dir_base + r"\OBM_error_log.xlsx"
    all_times_excel = save_dir_base + r"\OBM_raw_file_timestamps.xlsx"
    del_arr = [error_log_txt, error_log_xlsx, all_times_excel]
    for dir in del_arr:
        try:
            os.remove(dir)
        except FileNotFoundError:
            pass
    initialize_error_files(save_dir_base) 

def stitch_EDF_and_cut(e_file_directory, save_dir_base):
    """
    Process OBM EEG files by stitching segments and extracting analysis windows.
    
    Differs from Nicolet version by using timestamp-based pause detection rather
    than annotation parsing. Processes files in channel-specific subdirectories.
    
    Parameters:
    -----------
    e_file_directory : str
        Path to FIS directory (e.g., 'path/to/FIS123a/')
        Must contain channel subdirectory specified in OBM_channel config
    save_dir_base : str
        Base directory for output files and logs
        
    Returns:
    --------
    None or 0
        Returns 0 on error conditions (missing channel, corrupted files, etc.)
        Returns None on successful processing
        
    OBM-Specific Processing:
    -----------------------
    1. Navigate to channel subdirectory (e.g., FIS123a/CrossEeg/)
    2. Sort files using XXX_YYYY.edf naming convention
    3. Extract meas_date timestamps from EDF headers
    4. Calculate inter-file gaps as pause durations
    5. Zero-pad pause periods for continuous stream
    6. Process surgery/analysis windows based on config
    
    Pause Detection Algorithm:
    -------------------------
    For segment i > 1:
        pause_duration = meas_start[i] - meas_end[i-1]
        if pause_duration > 0:
            insert zeros for pause_duration seconds
            
    Error Conditions:
    ----------------
    - Channel directory missing: Channel not available for this FIS
    - Corrupted EDF files: Invalid or unreadable file format
    - Invalid timestamps: meas_date not datetime object
    - alex_raw conversion failure: Data processing error
    - Window building failure: Analysis window extraction error
    
    Side Effects:
    ------------
    - Creates comprehensive error logs
    - Updates Excel metadata files
    - Performs memory cleanup (garbage collection)
    """  
    import gc
    gc.collect()
    
    print("Stitching and cutting files.")
    channel = vars_dict["OBM_channel"]
    voltages_arr = [] #This will be an array of arrays, each of which contains voltage data from the pause or  the edf file
    times_arr = [] #At the end, srate will be used to build this out starting from 0 sec.
    paths_arr = [] #This will be used to verify that files were stitched or concatenated in the correct order

    final_file_name = str(os.path.basename(os.path.normpath(e_file_directory))) #This DOES NOT include .edf; e.g., will be FIS155b
    
    def custom_sort(file_list):
        """
        Sort OBM .edf files by 4-digit numeric suffix in XXX_YYYY.edf format.
        
        OBM files from EDFbrowser follow XXX_YYYY.edf naming where YYYY is a
        4-digit sequence number (e.g., recording_0001.edf, recording_0002.edf).
        Only files matching this pattern are included in processing.
        
        Parameters:
        -----------
        file_list : list of str
            List of .edf filenames from OBM directory
            
        Returns:
        --------
        list of str
            Filtered and sorted filenames in chronological order
            Only includes files with _YYYY.edf pattern
            
        Sorting Logic:
        -------------
        1. Filter: Only keep files matching _YYYY.edf pattern
        2. Extract: 4-digit YYYY suffix as integer
        3. Sort: By numeric value of YYYY suffix
        
        Example:
        --------
        Input: ['recording_0001.edf', 'recording_0010.edf', 'recording_0002.edf', 'other.edf']
        Output: ['recording_0001.edf', 'recording_0002.edf', 'recording_0010.edf']
        Note: 'other.edf' excluded (doesn't match pattern)
        """
        def sort_key(filename):
            # Extract the numeric suffix after the underscore
            match = re.search(r'_(\d{4})\.edf$', filename)
            if match:
                num_suffix = int(match.group(1))  # Extract YYYY as an integer
            else:
                num_suffix = None  # Exclude files that don't match the pattern
            return num_suffix
        # Filter out files without the _YYYY pattern
        filtered_files = [f for f in file_list if sort_key(f) is not None]
        # Sort the filtered list with the custom sort key
        return sorted(filtered_files, key=sort_key)
    
    # OBM files are organized in channel-specific subdirectories
    # Original path: FIS123a/
    # Target path: FIS123a/CrossEeg/ (or LeftEeg/, RightEeg/)
    e_file_dir_original = e_file_directory
    e_file_directory = f"{e_file_directory}\\{channel[0]}"

    if not os.path.exists(e_file_directory):
        add_error_to_excel("conversion", e_file_dir_original, save_dir_base)
        record_error(f"The file at {e_file_dir_original} did not have the channel we asked for (i.e., {channel[0]}.). Therefore, the file was skipped.", save_dir_base)
        record_file_info(e_file_directory, [0, 0], "N/a", save_dir_base, "-")
        return 0
    
    sorted_files = custom_sort(os.listdir(e_file_directory))
    pause_indices = []
    signal_length = 0
    pause_length_total = 0
    segment_i = 0
    meas_start_datetime_arr = []
    meas_end_datetime_arr = []


    for edf_file in sorted_files:
        if edf_file.endswith('.edf'):
            segment_i = segment_i + 1
            edf_path = os.path.join(e_file_directory, edf_file) #E.g., a6c6f55e-34aa-4280-8f73-bed63e619fb2FIS101post2\FIS101post22.edf
            #
            try:
                raw_EDF = import_edf_from_path(str(edf_path), downsample = vars_dict["downsample_rate"])
            except ValueError as e:
                error_str = f"Error importing file: {edf_path}. The specific error was {e}. Since this error was simply with importing the edf file, this was likely an issue with the file being corrupted. The file {edf_path} was NOT saved."
                add_error_to_excel("corrupt",e_file_directory, save_dir_base)
                record_error(error_str, save_dir_base)
                return 0 #We terminate these since we can't realistically stitch when one of the EDF files is corrupted
            # Extract measurement start time from EDF header
            # This is critical for OBM pause detection
            meas_start_datetime = raw_EDF.info["meas_date"]
            # Validate timestamp is proper datetime object
            if not isinstance(meas_start_datetime, datetime):
                add_error_to_excel("corrupt", edf_file, save_dir_base)
                record_error(f"The meas_date attribute of the info matrix in {edf_file} was not a datetime object.", save_dir_base)
                record_file_info(e_file_directory, [0, 0], alex_raw_stitched_and_cut.num_fis, save_dir_base, "-")
                return 0
            # Store for pause calculation with next segment
            meas_start_datetime_arr.append(meas_start_datetime)
            # Calculate end time for gap detection
            my_signal_length = raw_EDF.n_times/raw_EDF.info['sfreq'] #Length in seconds
            meas_end_datetime = meas_start_datetime + timedelta(seconds=my_signal_length)
            meas_end_datetime_arr.append(meas_end_datetime)


        else:
            continue #Since it's not an edf file (and therefore can't be processed here)
        pause_data = False #Turns true once we have data regarding a pause (i.e., will remain false when examining the first .edf in a .e folder, as this will have no pause/event information)
        paths_arr.append(edf_path)
        # OBM-specific pause detection using measurement timestamps
        # Unlike Nicolet (annotation-based), OBM uses EDF header timestamps
        if segment_i > 1:
            # Calculate gap between end of previous recording and start of current
            # meas_start_datetime_arr[i-1] = current file start time
            # meas_end_datetime_arr[i-2] = previous file end time
            pause_datetime = meas_start_datetime_arr[segment_i-1] - meas_end_datetime_arr[segment_i-2]
            # Get the time difference in seconds
            pause_length_seconds = pause_datetime.total_seconds()
            print(f"Pause found: {pause_datetime} equals {pause_length_seconds} seconds")
            pause_data = True
        elif segment_i == 1:
            pause_length_seconds = 0
        
        #The pause length (in SECONDS) is now stored in pause_length_seconds
        voltage_add_arr = []
        try:
            alex_raw_EDF = alex_raw(raw_EDF, which_channel=deepcopy(channel), filePath=str(edf_path), call_from="postop_file_finder")
        except ValueError as e:
            add_error_to_excel("conversion",e_file_directory, save_dir_base)
            record_error(f"Error converting the file {e_file_directory} to an alex_raw object. The issue specific error raised was: {e}. Therefore, the file was NOT saved.", save_dir_base)
            record_file_info(e_file_directory, [0, 0], alex_raw_stitched_and_cut.num_fis, save_dir_base, "-")
            return 0
        alex_raw_EDF.OBM = True
        sfreq = alex_raw_EDF.sampling_rate #(in Hz)
        n_data_points_to_add = int(sfreq * pause_length_seconds)
        edf_voltage_arr = np.round(alex_raw_EDF.EEG_data, decimals=2).astype(np.float32)
        voltage_add_arr = [0] * n_data_points_to_add #Fills empty space with zeros

        if pause_data: #This adds pause length of fillvar in those cases where there was a pause
            voltages_arr.extend(voltage_add_arr)
            pause_indices.append(len(voltages_arr)-1)
            voltages_arr.extend(edf_voltage_arr)
            print(f"Pause added to pause_indices. We have {len(pause_indices)} pauses recorded.")
            print(f"Stitched recording length thus far: {len(voltages_arr)/sfreq} sec.\n---")
            signal_length += len(edf_voltage_arr)
            pause_length_total += len(voltage_add_arr)
        else:
            voltages_arr.extend(edf_voltage_arr)
            print(f"Stitched recording length thus far: {len(voltages_arr)/sfreq} sec.\n---")
            signal_length += len(edf_voltage_arr)
    if segment_i == 0: #This means this directory had no .edf files, so we return 0 to break the function (which usually doesn't return anything)
        return 0

    
    print("Building arrays of all times. This may take a minute...")
    times_arr = np.arange(len(voltages_arr)) / sfreq
    print("Copying the last alex_raw_EDF object...")
    alex_raw_stitched_and_cut = alex_raw_EDF
    print("Copying voltages to the new alex_raw object...")
    alex_raw_stitched_and_cut.EEG_data = voltages_arr
    del voltages_arr
    print("Copying times to the new alex_raw object...")
    alex_raw_stitched_and_cut.EEG_times_from_zero = times_arr
    del times_arr
    alex_raw_stitched_and_cut.pause_indices_stitched_and_cut = pause_indices
    #-------------------------------------------------------------------------
    #Add datetime information to our alex_raw_stitched object
    chunks = vars_dict["windows"]
    if not (chunks == [["surgery"]] or chunks == [["CPB"]]):
        file_path = vars_dict["OBM_recording_time_data_path"]
        num_fis = int(alex_raw_stitched_and_cut.num_fis)
        print(f"Matching up num_fis: {num_fis}")

        print("Reading Excel for times...")
        df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl', dtype={
        'recording_start_date': str,
        'recording_start_time': str,
        'date_cx': str,
        'analyze_end_time': str})
        print("Extracting times of interest")

        #Get row of interest to match
        df['folder_name'] = df['folder_path'].apply(lambda x: x.split('\\')[-1])
        e_file_directory_name = e_file_directory.split('\\')[-2]
        rowi = df.index[df['folder_name'] == e_file_directory_name].tolist()[0]

        rec_start_date = str(df.at[rowi, 'recording_start_date']).zfill(8)
        try:
            rec_start_date = datetime.strptime(rec_start_date, '%d%m%Y').date()
        except ValueError as e:
            error_str= f"For the file {e_file_directory}, we got the error: {e} when reading the datetime."
            #print(error_str)
            add_error_to_excel("time",e_file_directory,save_dir_base)
            record_error(error_str, save_dir_base)
            record_file_info(e_file_directory, [0, 0], alex_raw_stitched_and_cut.num_fis, save_dir_base, "-")
            return 0
        rec_start_time = str(df.at[rowi, 'recording_start_time']).zfill(6)
        rec_start_time = datetime.strptime(rec_start_time, '%H%M%S').time()
        rec_start_date_time = datetime.combine(rec_start_date, rec_start_time)
        alex_raw_stitched_and_cut.rec_start_date_time = rec_start_date_time
        
        #Get date of surgery
        clindata_file_path = vars_dict["OBM_recording_time_data_path"]
        # Read the Excel sheet into a DataFrame
        df = pd.read_excel(clindata_file_path)
        df['fis_num'] = df['fis_num'].astype(int)
        # Find the row where 'fis_num' matches 'num_fis'
        row = df[df['fis_num'] == num_fis]
        if not row.empty:
            # Extract the value in the column 'date_cx'
            fecha_cx_str = row.iloc[0]['date_cx']
            # Convert the string to a datetime object
            fecha_cx_str = fecha_cx_str.strftime('%d/%m/%Y')  # Convert Timestamp to string
            cx_date = datetime.strptime(fecha_cx_str, '%d/%m/%Y')
        else:
            raise ValueError(f"No row for FIS{alex_raw_stitched_and_cut.fis} found in the sheet clinical_Data_Only.xlsx")

        if isinstance(df.at[rowi, 'analyze_end_time'], (int, float)) and np.isnan(df.at[rowi, 'analyze_end_time']):
            add_error_to_excel("time",e_file_directory,save_dir_base)
            record_error(f"There was no surgery time recorded in all_nicolet_folders_cleaning.xlsx for the file at {e_file_directory}.", save_dir_base)
            record_file_info(e_file_directory, [0, 0], alex_raw_stitched_and_cut.num_fis, save_dir_base, "-")
            return 0
        analyze_time = str(df.at[rowi, 'analyze_end_time'])
        analyze_time = analyze_time.split('.')[0]
        analyze_time = datetime.strptime(analyze_time, '%H%M').time()
        analyze_date_time = datetime.combine(cx_date, analyze_time)
        alex_raw_stitched_and_cut.rec_start_date_time = rec_start_date_time
        alex_raw_stitched_and_cut.cx_date = cx_date
        alex_raw_stitched_and_cut.analyze_end_date_time = analyze_date_time
    else:
        #Step 1: Access/extract information about the start of the recording
        file_path = vars_dict["OBM_recording_time_data_path"]
        num_fis = int(alex_raw_stitched_and_cut.num_fis)
        print(f"Matching up num_fis: {num_fis}")
        print("Reading Excel for times...")
        df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl', dtype={
        'recording_start_date': str,
        'recording_start_time': str,
        'date_cx': str})
        print("Extracting times of interest")
        df['folder_name'] = df['folder_path'].apply(lambda x: x.split('\\')[-1])
        e_file_directory_name = e_file_directory.split('\\')[-2]
        rowi = df.index[df['folder_name'] == e_file_directory_name].tolist()[0]
        rec_start_date = str(df.at[rowi, 'recording_start_date']).zfill(8)
        try:
            rec_start_date = datetime.strptime(rec_start_date, '%d%m%Y').date()
        except ValueError as e:
            error_str= f"For the file {e_file_directory}, we got the error: {e} when reading the datetime."
            #print(error_str)
            add_error_to_excel("time",e_file_directory,save_dir_base)
            record_error(error_str, save_dir_base)
            record_file_info(e_file_directory, [0, 0], alex_raw_stitched_and_cut.num_fis, save_dir_base, "-")
            return 0
        rec_start_time = str(df.at[rowi, 'recording_start_time']).zfill(6)
        rec_start_time = datetime.strptime(rec_start_time, '%H%M%S').time()
        rec_start_date_time = datetime.combine(rec_start_date, rec_start_time)
        alex_raw_stitched_and_cut.rec_start_date_time = rec_start_date_time

        #Step 2: Get the date of the surgery
        #Get date of surgery
        surgery_info_file_path = vars_dict["surgery_times_path"]
        # Read the Excel sheet into a DataFrame
        df_surg = pd.read_excel(surgery_info_file_path)
        df_surg['fis_num'] = df_surg['fis_num'].astype(int)
        # Find the row where 'fis_num' matches 'num_fis'
        row = df_surg[df_surg['fis_num'] == num_fis]
        if not row.empty:
            # Extract the value in the column 'Fecha Cx'
            fecha_cx_str = row.iloc[0]['date_surgery']
            # Convert the string to a datetime object
            fecha_cx_str = fecha_cx_str.strftime('%d/%m/%Y')  # Convert Timestamp to string
            cx_date = datetime.strptime(fecha_cx_str, '%d/%m/%Y')
        else:
            raise ValueError(f"No row for FIS{alex_raw_stitched_and_cut.fis} found in the sheet surgery_time_data.xlsx")
        
        time_end_val = row.iloc[0]['time_end_cx']
        time_start_val = row.iloc[0]['time_start_cx']

        if (isinstance(time_end_val, (int, float)) and np.isnan(time_end_val)) or (isinstance(time_start_val, (int, float)) and np.isnan(time_start_val)) or time_start_val == "-" or time_end_val == "-":
            add_error_to_excel("time",e_file_directory, save_dir_base)
            record_error(f"There was no surgery start or end time recorded in surgery_time_data.xlsx for the file at {e_file_directory}.", save_dir_base)
            record_file_info(e_file_directory, [0, 0], alex_raw_stitched_and_cut.num_fis, save_dir_base, "-")
            return 0
        
        #Step 3: set times
        if chunks == [["surgery"]]:
            analyze_time = str(time_end_val)
            analyze_time = analyze_time.split('.')[0]
            analyze_time = datetime.strptime(analyze_time, '%H%M').time()
            analyze_date_time = datetime.combine(cx_date, analyze_time)
            analyze_start_time = str(time_start_val)
            analyze_start_time = analyze_start_time.split('.')[0]
            analyze_start_time = datetime.strptime(analyze_start_time, '%H%M').time()
            analyze_start_date_time = datetime.combine(cx_date, analyze_start_time)
            alex_raw_stitched_and_cut.rec_start_date_time = rec_start_date_time
            alex_raw_stitched_and_cut.cx_date = cx_date
            alex_raw_stitched_and_cut.analyze_end_date_time = analyze_date_time
            alex_raw_stitched_and_cut.analyze_start_date_time = analyze_start_date_time
            alex_raw_stitched_and_cut.window_type = "postop"

    #Build EEG_times_from_analyze_end_time
    print("Building array of times relative to end of surgery...")
    rec_start_date_time = alex_raw_stitched_and_cut.rec_start_date_time
    analyze_end_date_time = alex_raw_stitched_and_cut.analyze_end_date_time
    sampling_rate = alex_raw_stitched_and_cut.sampling_rate
    eeg_times_length = len(alex_raw_stitched_and_cut.EEG_times_from_zero)
    initial_diff_seconds = (rec_start_date_time - analyze_end_date_time).total_seconds()

    time_deltas = np.arange(eeg_times_length) / sampling_rate
    seconds_from_end_date_time = initial_diff_seconds + time_deltas

    alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time = np.array(seconds_from_end_date_time) #NB: this is stored as SECONDS
    
    current_path = alex_raw_stitched_and_cut.file_path_current
    file_name_no_pkl = os.path.split(os.path.split(os.path.split(current_path)[0])[0])[1]
    new_file_name = file_name_no_pkl+"_stitched_cut"+'.pkl'
    new_path = os.path.join(save_dir_base + r"\stitched_cut_efficient", new_file_name)
    original_times = deepcopy([alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[0], alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[-1]])

    del time_deltas
    del seconds_from_end_date_time 

    pause_times_postop = []
    print(f"We have {len(pause_indices)} pauses recorded.")
    for pause_index in pause_indices:
        pause_times_postop.append(round(alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[pause_index] / 3600, 2))
    if pause_times_postop:
        print (f"The pauses occurred at the following number of hours postop: {pause_times_postop}\n\n")
    alex_raw_stitched_and_cut.pause_times_postop_hrs = pause_times_postop

    #-------------------------------------------------------------------------
    #Re-build our alex_raw.EEG_data, alex_raw.EEG_times_from_zero, and alex_raw.EEG_times_from_analyze_end_time arrays to contain only data in the windows of interest. Additionally, store the windows and a stitched_and_cut boolean.
    print("Cutting down data...")
    if not (chunks == [["surgery"]] or chunks == [["CPB"]]):
        try:
            window_indices, windows = alex_raw_stitched_and_cut.get_window_indices(chunks) #window_indices are the pairs of indices corresponding to each window in windows (e.g., if window_indices is [a,b], then alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[a] corresponds to the seconds values at 20 hr (if windows is [[20,24]]))
        except WindowBuildError as e:
            error_str= f"For the file {e_file_directory}, we got the error: {e} when trying to build windows."
            #print(error_str)
            add_error_to_excel("window",e_file_directory,save_dir_base)
            record_error(error_str, save_dir_base)
            record_file_info(e_file_directory, [0, 0], alex_raw_stitched_and_cut.num_fis, save_dir_base, "-")
            return 0
    else:
        windows = [[(alex_raw_stitched_and_cut.analyze_start_date_time - alex_raw_stitched_and_cut.analyze_end_date_time).total_seconds() / 3600, 0]]
        print(f"Our surgery times are {windows[0][0]}-{windows[0][1]} hrs.")
        try:
            window_indices, windows = alex_raw_stitched_and_cut.get_window_indices(windows) #window_indices are the pairs of indices corresponding to each window in windows (e.g., if window_indices is [a,b], then alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[a] corresponds to the seconds values at 20 hr (if windows is [[20,24]]))
        except WindowBuildError as e:
            error_str= f"For the file {e_file_directory}, we got the error: {e} when trying to build windows."
            #print(error_str)
            add_error_to_excel("window",e_file_directory,save_dir_base)
            record_error(error_str, save_dir_base)
            record_file_info(e_file_directory, [0, 0], alex_raw_stitched_and_cut.num_fis, save_dir_base, "-")
            return 0
    window_lengths_secs = []
    my_windows = []
    data_slices = []
    times_from_zero_slices = []
    times_analyze_slices = []
    i = 0
    for window in window_indices: #This is only for windows with data. NB: window here is a window index in alex_raw.EEG_data (or EEG_times_from_zero, etc)
        index1 = window[0]
        index2 = window[1]
        if (index2+1)<len(alex_raw_stitched_and_cut.EEG_data):
            data_slices.append(alex_raw_stitched_and_cut.EEG_data[index1:index2+1])
            times_from_zero_slices.append(alex_raw_stitched_and_cut.EEG_times_from_zero[index1:index2+1])
            times_analyze_slices.append(alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[index1:index2+1])
        else:
            data_slices.append(alex_raw_stitched_and_cut.EEG_data[index1:index2])
            times_from_zero_slices.append(alex_raw_stitched_and_cut.EEG_times_from_zero[index1:index2])
            times_analyze_slices.append(alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[index1:index2])
        my_windows.append(windows[i])
        window_length = times_from_zero_slices[i][-1] - times_from_zero_slices[i][0] + 1/alex_raw_stitched_and_cut.sampling_rate
        window_lengths_secs.append(window_length) #Each window 
        i = i + 1
    
    alex_raw_stitched_and_cut.EEG_data = np.concatenate(data_slices)
    del data_slices
    alex_raw_stitched_and_cut.EEG_times_from_zero = np.concatenate(times_from_zero_slices)
    del times_from_zero_slices
    alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time = np.concatenate(times_analyze_slices)
    del times_analyze_slices

    alex_raw_stitched_and_cut.windows = my_windows #This way the object only retains the window it contains
    alex_raw_stitched_and_cut.stitched_and_cut = True
    alex_raw_stitched_and_cut.window_indices, alex_raw_stitched_and_cut.windows = alex_raw_stitched_and_cut.get_window_indices(alex_raw_stitched_and_cut.windows) #Re-assigning after we've updated the data
    alex_raw_stitched_and_cut.window_lengths_secs = window_lengths_secs
        
    stitched_file_name=final_file_name #E.g., FIS101post2
    alex_raw_stitched_and_cut.saved_once = False
    alex_raw_stitched_and_cut.file_path_current = os.path.join(save_dir_base + r"\stitched_cut_full_files",stitched_file_name) #Because our file name is changing, we have to manually update self.file_path_current
    alex_raw_stitched_and_cut.add_change_log(f"Data was stitched together from the files {paths_arr} in the order shown. Pauses were retained, and their duration was marked with 0 in the object. Furthermore, the data was parsed to only include times from the windows: {windows}; of these, the windows that were found in this file were {alex_raw_stitched_and_cut.windows}, which was indeed in line with the Excel file.")
    print("Stitched file built. Now saving...")
    #alex_raw_stitched_and_cut.save_as(r"E:\EEG_Pipeline_Continuous\1_stitched_cut\stitched_cut_full_files","_stitched_cut")
    print("Creating alex_raw_efficient object...")
    alex_raw_small = alex_raw_efficient(alex_raw_stitched_and_cut)
    alex_raw_small.save_as_easy(save_dir_base + r"\stitched_cut_efficient","_stitched_cut")
    record_file_info(e_file_directory, original_times, alex_raw_stitched_and_cut.num_fis, save_dir_base, alex_raw_stitched_and_cut.windows[0], stitched_path = new_path)
    print("Moving on to next file...")


    return None #Done so local variables clear

def iterate_stitch_o(save_dir_base):
    r""""
    Function to iterate over OBM EDF+C files (these are made using EDFbrowser ahead of time). It decides whether or not to process a file based on whether or not the recording_time column in the Excel file E:\EEG_Pipeline_Continuous\1_stitched_cut\time_data\all_OBM_clean.xlsx has an entry. (I.e., if it does not have an entry, we process it).

    Inputs:
    - chunks (arr of int arr): an array of the time windows (expressed as two-entry arrays, where entry 1 is the start time of the chunk in HOURS, and entry 2 is the end time in HOURS). E.g., [[20, 24], [44, 48]]
    - clear (bool): if True, we will clear various Excel files related to the data processing as well as time data in E:\EEG_Pipeline_Continuous\1_stitched_cut\time_data\all_OBM_clean.xlsx. NB: this will not clear entries in the output folders for the stitched files
    - channel (str): the EEG channel from the OBM we would like to analyze. Options are "CrossEeg", "LeftEeg", and "RightEeg". Data is forthcoming regarding the specifics of the electrodes used in each channel.
    """
    clean_up_logs_and_excel(save_dir_base)
    OBM_folder = vars_dict["OBM_folder"]
    for subdirectory in os.listdir(OBM_folder):
        subdirectory_path = os.path.join(OBM_folder, subdirectory)
        stitch_EDF_and_cut(subdirectory_path, save_dir_base)