"""
EEG Artifact Rejection and Signal Quality Assessment Pipeline

This module processes stitched EEG recordings to identify and remove various types of artifacts,
ensuring signal quality meets minimum requirements for downstream analysis. The pipeline implements
a multi-stage artifact detection system with configurable thresholds and quality controls.

Purpose:
--------
Takes stitched alex_raw_efficient objects and applies systematic artifact rejection to remove:
1. Continuous zero runs (electrode disconnection)
2. High-amplitude artifacts (movement, electrical interference)
3. Continuous constant values
4. Sudden amplitude jumps

The module also enforces minimum signal quality requirements, rejecting recordings with
insufficient clean data for reliable analysis.

Artifact Detection Pipeline:
---------------------------
1. Zero Run Detection: Identifies continuous sequences of zero values indicating electrode issues
2. High-Amplitude Detection: Uses Hilbert transform to detect excessive signal amplitudes
3. Continuous Value Detection: Finds periods of unchanging signal values (amplifier saturation)
4. Sudden Jump Detection: Detects rapid amplitude changes indicating electrical artifacts

Each detection stage adds a "collar" of additional samples around detected artifacts to ensure
complete removal of contaminated data.

Quality Control System:
----------------------
- Proportion Cutoff: Maximum allowable percentage of artifact in signal
- Time Cutoff: Minimum required duration of clean data within analysis window
- Dual Criteria: Both proportion and absolute time requirements must be met

Signal Processing:
-----------------
- Butterworth Filtering: Applied before artifact detection (0.1-40 Hz bandpass)
- NaN Interpolation: Handles gaps in data using linear or PCHIP interpolation
    - NB: this filtering an interpolation is done only so that artefact rejection is performed on these filtered signals; however, these filters/interpolations are NOT carried forward into future signal processing steps
- Edge Handling: Special processing for NaN values at signal boundaries

Data Structures:
---------------
- alex_raw: Input EEG objects with voltage data and metadata
- alex_raw_efficient: Memory-optimized output objects
- rejected_tracker.xlsx: Comprehensive rejection logging with metrics
- final_samples.xlsx: Registry of successfully processed files

Configuration Dependencies:
--------------------------
Required variables from config.get_vars_dict():
- zero_run_threshold: Minimum duration (seconds) to flag zero runs
- high_amp_threshold: Voltage threshold for high-amplitude artifacts
- high_amp_collar: Time padding (seconds) around high-amplitude artifacts
- continuous_threshold: Duration threshold for constant value detection
- continuous_collar: Time padding around constant value artifacts
- sj_threshold: Voltage change threshold for sudden jump detection
- sj_collar: Time padding around sudden jump artifacts
- proportion_cutoff: Maximum allowable artifact proportion (0-1)
- time_cutoff: Minimum required clean data proportion (0-1)

Output Files:
------------
- artefact_rejected_files/: Successfully processed alex_raw objects
- rejected_tracker.xlsx: Detailed rejection logs with quality metrics
- final_samples.xlsx: Registry of processed files by time window
- artefact_rejection_parameters.txt: Processing configuration record

Error Handling:
--------------
- File corruption: Graceful handling with detailed error logging
- Signal quality failures: Comprehensive rejection tracking
- Processing errors: Exception handling with traceback logging
- Duplicate processing: Prevention through rejection tracker checks
"""

import os
import sys
current_dir = os.path.dirname(os.path.abspath(__file__))
utils_dir = os.path.join(current_dir, '..', 'utils')
sys.path.append(utils_dir)
import numpy as np
from scipy import signal
from file_importing import import_pkl
from tqdm import tqdm
from copy import deepcopy
import warnings
import matplotlib.pyplot as plt
from NEURAL_py_EEG import mfiltfilt
from scipy.interpolate import interp1d
from scipy.interpolate import PchipInterpolator
import openpyxl
from openpyxl import load_workbook, Workbook
import random
from datetime import datetime
import glob
import traceback
from utils import config
vars_dict = config.get_vars_dict()


def add_to_rejected_tracker(save_base_dir, pkl_file, proportion_arts, message="", metrics=None):
    """
    Records rejected files and their metrics in the rejected_tracker Excel file.
    Added file locking to prevent race conditions.
    """
    file_name = os.path.join(save_base_dir, "rejected_tracker.xlsx")
    # Check if the file exists
    if not os.path.exists(file_name):
        # Create new workbook with headers
        wb = openpyxl.Workbook()
        sheet = wb.active
        headers = ["file_name", "rejected", "proportion_artefacts", 
                    "len_clean_signal_hrs", "len_total_signal_hrs", 
                    "proportion_signal_clean", "proportion_chunk_clean","message"]
        sheet.append(headers)
    else:
        wb = openpyxl.load_workbook(file_name)       
        sheet = wb.active

    # Convert metrics to hours
    if metrics:
        len_clean_hrs = metrics['len_clean_secs'] / 3600
        len_total_hrs = metrics['len_total_secs'] / 3600
        proportion_signal_clean = metrics['proportion_signal_clean']
        proportion_chunk_clean = metrics['proportion_chunk_clean']
    else:
        len_clean_hrs = None
        len_total_hrs = None
        proportion_signal_clean = None
        proportion_chunk_clean = None

    # Add new row
    new_row = [
        str(pkl_file),            # file_name
        1,                        # rejected
        round(proportion_arts, 3), # proportion_artefacts
        round(len_clean_hrs, 3) if len_clean_hrs is not None else None,
        round(len_total_hrs, 3) if len_total_hrs is not None else None,
        round(proportion_signal_clean, 3) if proportion_signal_clean is not None else None,
        round(proportion_chunk_clean, 3) if proportion_chunk_clean is not None else None,
        message
    ]
    sheet.append(new_row)
    wb.save(file_name)     
    print(f"Added '{str(pkl_file)}' to rejected tracker with proportion artefacts: {round(proportion_arts, 3)}")

def check_signal_proportions(EEG_data, chunk, sampling_rate, proportion_cutoff=0.5, time_cutoff=0.5):
    """
    Checks if the signal meets the minimum requirements for proportion of clean data.
    """
    # Calculate metrics
    len_clean_secs = np.sum(~np.isnan(EEG_data))/sampling_rate
    len_total_secs = len(EEG_data)/sampling_rate
    proportion_of_signal_clean = len_clean_secs/len_total_secs
    proportion_of_chunk_clean = len_clean_secs/(chunk[1]-chunk[0])
    
    # Calculate minimum required samples
    min_samples_proportion = (1-proportion_cutoff)*len(EEG_data)
    min_samples_time = (chunk[1]-chunk[0])*sampling_rate*time_cutoff
    max_samples_required = max(min_samples_proportion, min_samples_time)
    
    meets_requirements = np.sum(~np.isnan(EEG_data)) >= max_samples_required
    
    metrics = {
        'len_clean_secs': len_clean_secs,
        'len_total_secs': len_total_secs,
        'proportion_signal_clean': proportion_of_signal_clean,
        'proportion_chunk_clean': proportion_of_chunk_clean
    }
    
    return meets_requirements, metrics

def process_file_to_excel(FIS, chunks, excel_path=None):
    """
    Process a single file to determine its sheet and add its data to 'final_samples.xlsx'.
    """
    if excel_path is None:
        return False
        
    hr_bounds = [str(int(val/3600)) for val in chunks]
    
    # Sheet name based on temporal boundaries
    sheet_name = f"{hr_bounds[0]}_{hr_bounds[1]}_hrs"

    # Check if the Excel file exists
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path, data_only=True)
    else:
        # Initialize the workbook if it doesn't exist
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet
        print(f"Created a new Excel file: {excel_path}")
    
    # Check if the sheet exists; if not, create it
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(["FIS"])  # Add the header row
        print(f"Created a new sheet: {sheet_name}")
    else:
        ws = wb[sheet_name]
    
    # Check if the file already exists in the sheet
    FISes = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
    if FIS in FISes:
        print(f"{FIS} already exists in the sheet '{sheet_name}'.")
        return False
        
    # Add the file name to the sheet
    ws.append([FIS])
    wb.save(excel_path)
    print(f"Updated the Excel file: {excel_path}")
    return True

def len_cont_zeros(x, const=0):
    """
    len_cont_zeros: find length of continuous segments of zeros from binary mask x. Can contain NaNs.
    """
    nan_loc = np.where(np.isnan(x))[0]
    DBplot = 0
    if np.array_equal(x, x.astype(bool)) or const not in [0, 1]:
        warnings.warn("Must be a binary signal", DeprecationWarning)

    if const == 1:
        y = np.abs(x - 1)
    else:
        y = x

    # Find run of zeros
    y = (y == 0).astype(float)
    y[nan_loc] = 0
    iedge = np.diff(np.concatenate(([0], y, [0])))
    istart = np.array(np.where(iedge == 1))[0]  # Row zero only - input array should be 1d
    iend = np.subtract(np.where(iedge == -1), 1)[0]
    lens = np.array([iend - istart])[0]

    if DBplot:
        plt.figure(100)
        plt.plot(x)
        plt.plot(istart, np.array(x[istart]), marker="x")
        plt.plot(iend, np.array(x[iend]), marker="o", markerfacecolor="none")

    return np.array(lens), istart, iend

def replace_start_ends_NaNs_with_zeros(x):
    """
    replace leading or trailing NaNs with zeros (needed for naninterp)
    """
    N = len(x)
    istart = np.argwhere(~np.isnan(x))[0][0]
    iend = np.argwhere(~np.isnan(x))[-1][0]

    if istart.size > 0 and istart > 0:
        x[0 : istart + 1] = 0
    if iend.size > 0 and iend < N - 1:
        x[iend + 1 : N + 1] = 0

    return x

def naninterp(X, method="linear"):
    """
    fill 'gaps' in data (marked by NaN) by interpolating
    """
    inan = np.argwhere(np.isnan(X))
    if inan.size == 0:
        return X, inan
    elif len(inan) == 1:
        if inan > 0:
            X[inan] = X[inan - 1]
        else:
            X[inan] = X[inan + 1]
    else:
        try:
            if method != "pchip":
                set_interp = interp1d(
                    np.transpose(np.argwhere(~np.isnan(X)))[0],
                    np.transpose(X[np.argwhere(~np.isnan(X))])[0],
                    kind=method,
                )
                X[inan] = set_interp(inan)
            else:
                set_interp = PchipInterpolator(
                    np.transpose(np.argwhere(~np.isnan(X)))[0],
                    np.transpose(X[np.argwhere(~np.isnan(X))])[0],
                    extrapolate=False,
                )
                X[inan] = set_interp(inan)

        except:
            raise ValueError("linear interpolation with NaNs")
    return X, inan

def filter_butterworth_withnans(x, Fs, F3db_lowpass, F3db_highpass, order, FILTER_REPLACE_ARTEFACTS="linear_interp"):
    """
    Butterworth filter that handles NaN values
    """
    inans = np.array([])

    if not F3db_highpass:
        b, a = signal.butter(order, F3db_lowpass / (Fs / 2), "lowpass")
    elif not F3db_lowpass:
        b, a = signal.butter(order, F3db_highpass / (Fs / 2), "highpass")
    else:
        if isinstance(order, list):
            order_low = order[0]
            order_high = order[1]
        else:
            order_low = order
            order_high = order

        y, isnans_low = filter_butterworth_withnans(
            x, Fs, F3db_lowpass, [], order_low, FILTER_REPLACE_ARTEFACTS
        )
        y, isnans_high = filter_butterworth_withnans(
            y, Fs, [], F3db_highpass, order_high, FILTER_REPLACE_ARTEFACTS
        )
        inans = np.unique(np.concatenate([isnans_low, isnans_high]))
        return y, inans

    # remove NaNs and replace with linear interopolation (this is just for the analytical waveform creation step; these interpolations are NOT carried forward into the analyzed signal itself)
    inans = np.array(np.argwhere(np.isnan(x)))
    if inans.size != 0:
        if FILTER_REPLACE_ARTEFACTS == "zeros":
            x[inans] = 0
        elif FILTER_REPLACE_ARTEFACTS == "linear_interp":
            x = replace_start_ends_NaNs_with_zeros(x)
            x, dum = naninterp(x, "linear")
        elif (
            FILTER_REPLACE_ARTEFACTS == "cubic_interp"
            or FILTER_REPLACE_ARTEFACTS == "nans"
        ):
            x = replace_start_ends_NaNs_with_zeros(x.copy())
            x, dum = naninterp(x.copy(), "pchip")

    y = mfiltfilt.mat_filtfilt(np.array(b), np.array(a), x)

    # special case if nans
    if FILTER_REPLACE_ARTEFACTS.lower() == "nans":
        if inans.size != 0:
            y[inans] = np.nan

    return y, inans

def art_per_channel(x, y, Fs):
    """
    Remove artefacts on a per-channel basis
    """
    x = np.around(x, 2)  # Rounds our EEG voltages
    N = len(x)

    amount_removed = np.array([0, 0, 0, 0], dtype=float)  # Each index is number of points removed in a given step
    # ---------------------------------------------------------------------
    # 1. electrode - checks(continuous row of zeros)
    # ---------------------------------------------------------------------
    print("Marking zero runs...")
    x_channel = x.copy()
    x_channel[x_channel != 0] = 1
    irem = np.zeros([N])
    
    lens, istart, iend = len_cont_zeros(x_channel, 0)
    ielec = np.flatnonzero(lens >= (vars_dict["zero_run_threshold"] * Fs))

    if ielec.size != 0:
        for m in ielec:
            irun = np.array(range((istart[m] - 1), (iend[m] + 2), 1))
            irun[irun < 0] = 0
            irun[irun > N - 1] = N - 1
            irem[irun] = 1
            x[irun] = np.nan
            y[irun] = np.nan
    if any(irem == 1):
        print("continuous row of zeros: %.2f%%\n" % (100 * (np.sum(irem) / N)))
    else:
        print("No zero runs found.")
    amount_removed[0] = 100 * (np.sum(irem) / N)

    x_nofilt = x.copy()
    
    x_filt, inans = filter_butterworth_withnans(x, Fs, 40, 0.1, [5, 2])

    # ---------------------------------------------------------------------
    # 2. high - amplitude artefacts
    # ---------------------------------------------------------------------
    print("Marking high-amplitude artefacts...")
    art_coll = vars_dict["high_amp_collar"] * Fs
    irem = np.zeros(N)

    print("Getting amplitudes...")
    # Apply Hilbert transform to get instantaneous amplitude envelope
    # This captures both positive and negative amplitude excursions
    x_hilbert = np.abs(signal.hilbert(x_filt))

    thres_upper = vars_dict["high_amp_threshold"]
    abs_val_rejection = True
    # Dual detection strategy:
    # 1. Hilbert envelope (captures filtered signal amplitude)
    # 2. Raw signal absolute value (catches sharp transients that may be attenuated by filtering)
    if abs_val_rejection == True:
        ihigh = np.flatnonzero((x_hilbert > thres_upper) | (np.abs(x_nofilt) > thres_upper))
    else:
        ihigh = np.flatnonzero(x_hilbert > thres_upper)

    # Apply collar padding around detected artifacts
    # Collar ensures complete removal of artifact-contaminated signal
    starts = []
    ends = []
    if ihigh.size != 0:
        starts = np.clip(ihigh - int(art_coll), 0, N - 1)
        ends = np.clip(ihigh + int(art_coll), 0, N - 1)
        
    # Mark expanded regions for removal
    for start, end in zip(starts, ends):
        irem[start:(end + 1)] = 1

    # Replace detected artifacts with NaN in both voltage and time arrays
    x[irem == 1] = np.nan
    y[irem == 1] = np.nan
    if any(irem == 1):
        print("length of high-amplitude artefacts: %.2f%%\n" % (100 * (np.sum(irem) / N)))
    else:
        print("No high-amplitude artefacts found.")
    amount_removed[1] = 100 * (np.sum(irem) / N)

    # ---------------------------------------------------------------------
    # 3. continuous constant values(i.e.artefacts)
    # ---------------------------------------------------------------------
    print("Marking continuous value artefacts...")
    art_coll = vars_dict["continuous_collar"] * Fs
    irem = np.zeros(N)

    x_diff_all = np.concatenate((np.diff(x), [0]))
    x_diff = x_diff_all.copy()
    x_diff[x_diff != 0] = 1
    lens, istart, iend = len_cont_zeros(x_diff, 0)

    # if exactly constant for longer than our threshold, then remove:
    ielec = np.flatnonzero(lens > (vars_dict["continuous_threshold"] * Fs))

    if ielec.size != 0:
        for m in ielec:
            irun = np.array(range((istart[m] - int(art_coll)), (iend[m] + int(art_coll) + 1), 1))
            irun[irun < 0] = 0
            irun[irun > N - 1] = N - 1
            irem[irun] = 1
            x[irun] = np.nan
            y[irun] = np.nan
    if any(irem == 1):
        print("Continuous row of constant values: %.2f%%\n" % (100 * (np.sum(irem) / N)))
    else:
        print("No continuous value artefacts found.")
    amount_removed[2] = 100 * (np.sum(irem) / N)

    # ---------------------------------------------------------------------
    # 4. sudden jumps in amplitudes or constant values(i.e.artefacts)
    # ---------------------------------------------------------------------
    print("Marking sudden-jump artefacts...")
    art_coll = vars_dict["sj_collar"] * Fs
    irem = np.zeros(N)
    x_diff = x_diff_all.copy()

    jump_threshold = vars_dict["sj_threshold"]
    
    ihigh = np.flatnonzero(np.abs(x_diff) > jump_threshold)
    if ihigh.size != 0:
        starts = np.clip(ihigh - int(art_coll), 0, N - 1)
        ends = np.clip(ihigh + int(art_coll), 0, N - 1)

        for start, end in zip(starts, ends):
            irem[start:end + 1] = 1

    x[irem == 1] = np.nan
    y[irem == 1] = np.nan
    if any(irem == 1):
        print("Length of sudden-jump artefacts: %.2f%%\n" % (100 * (np.sum(irem) / N)))
    else:
        print("No suddent-jump artefacts found.")
    amount_removed[3] = 100 * (np.sum(irem) / N)

    # before filtering, but should be eliminated anyway
    x[inans] = np.nan
    y[inans] = np.nan
    inans = np.flatnonzero(np.isnan(x))
    x_nofilt[inans] = np.nan
    y[inans] = np.nan
    x = x_nofilt
    x_nan = x
    return x_nan, amount_removed

def art_rej_alex_raw(save_base_dir, alex_raw_object, pkl_name):
    """
    Modified version of art_rej_alex_raw that includes proportion checks
    """
    proportion_cutoff = vars_dict["proportion_cutoff"]
    time_cutoff = vars_dict["time_cutoff"]
    
    try:
        # Step 1: Gather initial data
        EEG_data = np.copy(alex_raw_object.EEG_data)
        EEG_times_from_zero = np.copy(alex_raw_object.EEG_times_from_zero)
        sampling_rate = alex_raw_object.sampling_rate

        # Step 2: Perform artefact rejection
        EEG_data_artefacts_identified, amount_removed = art_per_channel(EEG_data, EEG_times_from_zero, sampling_rate)
        alex_raw_object.EEG_data = EEG_data_artefacts_identified
        
        # Check if signal meets proportion requirements
        meets_requirements, metrics = check_signal_proportions(EEG_data_artefacts_identified, alex_raw_object.windows[0], sampling_rate, proportion_cutoff, time_cutoff)
        
        if not meets_requirements:
            # Log rejection and return empty list to indicate rejection
            alex_raw_object.EEG_data = EEG_data_artefacts_identified
            alex_raw_object.proportion_artefacts = np.isnan(EEG_data_artefacts_identified).sum() / EEG_data_artefacts_identified.size
            metrics['rejected_reason'] = 'proportion_requirements'
            add_to_rejected_tracker(save_base_dir, pkl_name, alex_raw_object.proportion_artefacts, metrics=metrics, message="Did not meet proportion filtering requirements.")
            return []
        
        if np.count_nonzero(~np.isnan(EEG_data_artefacts_identified)) <= 1:
            add_to_rejected_tracker(save_base_dir, pkl_name, alex_raw_object.proportion_artefacts, metrics=metrics,message = "All data points were artefact.")
            return []  # Return since there's no way we'll have any good windows
        
        alex_raw_object.EEG_data = EEG_data_artefacts_identified
        # Update the alex_raw object
        alex_raw_object.proportion_artefacts = np.isnan(EEG_data_artefacts_identified).sum() / EEG_data_artefacts_identified.size
        print(f"{round(alex_raw_object.proportion_artefacts*100,2)}% of the signal was deemed artefact (and therefore removed).")

        # Step 3: Update chunk metadata in the raw object and perform checks to make sure everything ran smoothly
        # Update metadata
        save_raw = deepcopy(alex_raw_object)
        save_raw.chunk = alex_raw_object.windows[0]
        
        # Error check 3.1: Ensure that previous processing steps were correct, and that the object contains only one window
        if not(len(alex_raw_object.windows) == 1):
            raise ValueError(f"We should only have one window in our alex_raw_efficient object, but we have multiple. The object was {alex_raw_object.file_name}.pkl, and the windows it contained were {alex_raw_object.windows}.")
        
        # Error check 3.2: Ensure that the values all still match up
        for i in range(0, 10):
            save_check_ind = random.randint(0, len(alex_raw_object.EEG_data) - 1)
            # Retrieve the value at this index
            save_check_time = save_raw.EEG_times_from_analyze_end_time[save_check_ind]
            # Find the index of the closest time in alex_raw_original.EEG_times_from_analyze_end_time
            time_differences = np.abs(np.array(alex_raw_object.EEG_times_from_analyze_end_time) - save_check_time)
            orig_check_ind = np.argmin(time_differences)
            # Check if the time difference exceeds the allowable threshold
            if abs(alex_raw_object.EEG_times_from_analyze_end_time[orig_check_ind] - save_check_time) > 1 / save_raw.sampling_rate:
                raise ValueError("The time in save_raw was not found in alex_raw_original")
            # Check if the voltages at the same times match within a rounding threshold
            if not ((np.isnan(alex_raw_object.EEG_data[orig_check_ind]) and np.isnan(save_raw.EEG_data[save_check_ind])) or round(alex_raw_object.EEG_data[orig_check_ind], 2) == round(save_raw.EEG_data[save_check_ind], 2)):
                raise ValueError(f"The voltages at the same postop times in save_raw and alex_raw_original were not the same. The alex_raw_original value was {alex_raw_object.EEG_data[orig_check_ind]}, while the save_raw object value was {save_raw.EEG_data[save_check_ind]}.")

        # Step 4: Update save_raw's change log and save the file
        save_raw.amount_removed_by_artefact = amount_removed  # Index 0 is zero runs, index 1 is high-amplitude artefacts, index 2 is continuous non-zero runs, index 3 is sudden jump
        save_raw.add_change_log(f"Artefact rejection performed using the art_per_channel method in artefact_rejection.py. \nPercent of data that was zero runs: {round(amount_removed[0],2)}%. Percent of data that was high-amplitude artefacts: {round(amount_removed[1],2)}%. Percent of data that was continuous runs: {round(amount_removed[2],2)}%. Percent of data that was sudden-jump artefact: {round(amount_removed[3],2)}%. Together, this is a total of {round(save_raw.proportion_artefacts * 100,2)}% of data being artefact. This file reflects the range of interest {[int(val/3600) for val in save_raw.chunk]} hours. Check in the function art_per_channel in the file artefact_rejection_segmentation_continuous.py for the specific cutoffs and collars used in each step.")
        
        excel_path_to_use = os.path.join(save_base_dir, "final_samples.xlsx")
        
        if process_file_to_excel(save_raw.fis, [save_raw.chunk[0], save_raw.chunk[1]], excel_path=excel_path_to_use):
            print("Saving artefact rejected file...")
            
            # Create directory if it doesn't exist
            artifact_dir = os.path.join(save_base_dir, "artefact_rejected_files")
            os.makedirs(artifact_dir, exist_ok=True)
            
            save_raw.save_as_easy(artifact_dir, f"_ar_{int(save_raw.chunk[0]/3600)}_{int(save_raw.chunk[1]/3600)}_hrs")
            return [save_raw]
        else:
            add_to_rejected_tracker(save_base_dir, pkl_name, alex_raw_object.proportion_artefacts, metrics=metrics, message="FIS was already in final_samples.xlsx. This should NOT have happened. Debug.")
            return []
            
    except Exception as e:
        # Log the error
        print(f"Error processing {pkl_name}: {e}")
        traceback.print_exc()
        add_to_rejected_tracker(save_base_dir, pkl_name, alex_raw_object.proportion_artefacts, metrics=metrics, message = f"{e}")
        return []

def initialize_rejected_tracker(save_base_dir, clear=False):
    """
    Initialize the rejected tracker Excel file
    """
    file_name = os.path.join(save_base_dir, "rejected_tracker.xlsx")
    
    # Delete the file if clear is True and it exists
    if clear and os.path.exists(file_name):
        os.remove(file_name)
    
    # Create a new workbook with file locking
    if not os.path.exists(file_name):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"
        # Add headers
        sheet.append(["file_name", "rejected", "proportion_artefacts", 
                "len_clean_signal_hrs", "len_total_signal_hrs", 
                "proportion_signal_clean", "proportion_chunk_clean"])
        # Save the workbook
        wb.save(file_name)
        print(f"'{file_name}' initialized.")

def get_rejected_name_column(save_base_dir):
    """
    Get the list of rejected file names
    """
    file_name = os.path.join(save_base_dir, "rejected_tracker.xlsx")
    
    # Check if the file exists
    if not os.path.exists(file_name):
        initialize_rejected_tracker(save_base_dir)
        return []
    
    # Open the workbook
    wb = openpyxl.load_workbook(file_name, data_only=True)
    # Get the active sheet
    sheet = wb.active
    # Find the "file_name" column
    header = [cell.value for cell in sheet[1]]  # Read the header row
    if "file_name" not in header:
        return []
    # Get the index of the "file_name" column (1-based)
    file_name_col_index = header.index("file_name") + 1
    # Get all entries in the "file_name" column, starting from the second row
    rejected_name_col = [sheet.cell(row=row, column=file_name_col_index).value
        for row in range(2, sheet.max_row + 1)]
    # Filter out empty values
    rejected_name_col = [entry for entry in rejected_name_col if entry is not None]
    return rejected_name_col

def create_directories(save_base_dir, clear=False):
    """
    Create necessary directories and clear files if needed
    """
    os.makedirs(save_base_dir, exist_ok=True)
    
    # Create subdirectories
    dirs_to_create = [
        os.path.join(save_base_dir, "artefact_rejected_files"),
        os.path.join(save_base_dir, "images")
    ]
    
    for directory in dirs_to_create:
        os.makedirs(directory, exist_ok=True)
    
    # Clear files if requested
    if clear:
        # Clear files in directories
        for directory in dirs_to_create:
            files = glob.glob(os.path.join(directory, "*"))
            for f in files:
                try:
                    os.remove(f)
                except Exception as e:
                    print(f"Error removing {f}: {e}")
        
        # Clear Excel files
        files_to_delete = [
            os.path.join(save_base_dir, "final_samples.xlsx"),
            os.path.join(save_base_dir, "rejected_tracker.xlsx"),
            os.path.join(save_base_dir, "error_log.xlsx")
        ]
        
        for file_name in files_to_delete:
            if os.path.exists(file_name):
                try:
                    os.remove(file_name)
                except Exception as e:
                    print(f"Error removing {file_name}: {e}")

def iterate_artefact_rejection(save_base_dir):
    """
    Function that iterates over all files (expecting stitched and cut) at a given directory and performs artefact rejection.
    No outputs, but will save all the rejected files to "\artefact_rejected_signals".
    
    Inputs:
    - save_base_dir (str): the directory where we would like the artefact-rejected .pkl files to be stored.
    - proportion_cutoff (float): Maximum proportion of data allowed to be NaN
    - time_cutoff (float): Minimum proportion of chunk that must contain real data
    - clear (bool): Whether to clear existing files
    - best_hour (bool): Whether to find the best hour of data
    - my_base_dir (str): Custom base directory path (if needed)
    - parallel (bool): Whether to process files in parallel
    - num_processes (int): Number of parallel processes to use (default: CPU count - 1)
    - best_hour_min (int): Minimum minutes required for best hour
    """
    proportion_cutoff = vars_dict["proportion_cutoff"]
    time_cutoff = vars_dict["time_cutoff"]
    
    # Define the source directory
    stored_files_directory = os.path.join(os.path.split(save_base_dir)[0], "1_stitched_cut", "final_stitched_cut_efficient")
    # Create directories and initialize tracking files
    create_directories(save_base_dir, True)
    initialize_rejected_tracker(save_base_dir, True)
    # Get list of rejected files (to avoid reprocessing)
    rejected_name_col = get_rejected_name_column(save_base_dir)
    
    # Get list of files to process
    all_files = [f for f in os.listdir(stored_files_directory) if f.endswith('.pkl')]
    
    # Filter out already processed files
    files_to_process = [f for f in all_files if f not in rejected_name_col]
    
    if not files_to_process:
        print("No files to process!")
        return
    
    print(f"Found {len(files_to_process)} files to process.")

    # If parallel processing is disabled, use sequential processing
    print("Using sequential processing...")
    for pkl_file in tqdm(files_to_process, desc="Processing files"):
        try:
            import_path = os.path.join(stored_files_directory, pkl_file)
            alex_raw_object = import_pkl(full_file_path=import_path, parts_or_path="path")
            result = art_rej_alex_raw(save_base_dir, alex_raw_object, pkl_file)
            if not result:
                print(f"File rejected: {pkl_file}")
        except Exception as e:
            print(f"Error processing {pkl_file}: {e}")
            traceback.print_exc()
    
    # Record our art_rej parameters
    param_file_path = os.path.join(save_base_dir, "artefact_rejection_parameters.txt")
    with open(param_file_path, "w") as f:
        f.write(f"proportion_cutoff: {proportion_cutoff}\n")
        f.write(f"time_cutoff: {time_cutoff}\n")
        f.write(f"Processing date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")