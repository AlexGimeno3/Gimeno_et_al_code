"""
Alex Gimeno, 2025

Nicolet EEG Data Stitching and Processing Pipeline

This module processes EEG recordings exported from Nicolet machines, converting segmented .edf files
into unified, analyzable datasets. The Nicolet machine exports .e files as multiple .edf segments
based on recording events (pauses, stops), and this pipeline reconstructs the continuous signal.

Purpose:
--------
Takes a FIS (patient) folder containing multiple .edf files and combines them into a single
alex_raw_efficient object containing all voltage data in one continuous stream. The resulting
object is saved to the final_stitched_cut_efficient directory for downstream analysis.

Key Data Structures:
-------------------
1. alex_raw: Base EEG data object containing voltage data, timestamps, and metadata
2. alex_raw_efficient: Memory-optimized version of alex_raw for storage/transfer
3. voltages_arr: List accumulating voltage data from all .edf segments
4. times_arr: Numpy array of timestamps relative to recording start (seconds)
5. pause_indices: List tracking where recording pauses occurred in the voltage stream

Processing Workflow:
-------------------
1. Read and sort .edf files chronologically from FIS folder
2. Extract pause duration from .edf annotations (stored as "HH:MM:SS" format)
3. Insert zero-filled arrays during pause periods (for later artifact removal)
4. Concatenate all voltage data into continuous stream
5. Generate timestamp arrays relative to recording start and surgery end
6. Extract only data within specified time windows of interest
7. Save as memory-efficient object for analysis pipeline

Data Handling:
--------------
- Pauses between recordings: Filled with zeros (flatline for artifact rejection)
- Sampling rate: Preserved from original .edf files (typically downsampled)
- Channel selection: Configurable via nicolet_channels in config
- Time windows: Configurable analysis periods (e.g., 20-24 hours post-surgery)
- Memory management: Aggressive cleanup of large arrays during processing

Error Handling:
--------------
- Corrupted .edf files: Logged and skipped
- Missing timestamp data: Recorded in error logs
- Window building failures: Graceful failure with error logging
- File conversion errors: Comprehensive error tracking in Excel logs

Dependencies:
------------
- alex_raw/alex_raw_efficient: Custom EEG data objects
- file_importing: EDF file reading utilities
- WindowBuildError: Custom exception for time window issues
- config: Configuration management for paths and parameters

Output Files:
------------
- stitched_cut_efficient/: Processed alex_raw_efficient objects
- nicolet_raw_file_timestamps.xlsx: Processing metadata and timing info
- nicolet_error_log.xlsx/.txt: Comprehensive error tracking

Required configuration variables from config.get_vars_dict():
------------
- nicolet_channels: Channel selection for EEG data
- windows: Time windows for analysis (e.g., [[20,24]] for 20-24h post-surgery)  
- nicolet_recording_time_data_path: Excel file with recording metadata
- surgery_times_path: Excel file with surgery timing data
- downsample_rate: Downsampling factor for EEG data
- nicolet_folder: Base directory containing FIS folders

Expected directory structure:
------------
nicolet_folder/
  ├── FIS001post1/
  │   ├── recording.edf      # Base recording (no numeric suffix)
  │   ├── recording1.edf     # Continuation files (numbered)
  │   └── recording2.edf
  └── FIS002post2/
      └── ...

.edf annotation format for pauses:
------------
- Description contains "HH:MM:SS" format for pause duration
- Multiple descriptions possible but only one should contain time format
"""





from copy import deepcopy
import os
import sys
current_dir = os.path.dirname(os.path.abspath(__file__))
utils_dir = os.path.join(current_dir, '..', 'utils')
sys.path.append(utils_dir)
from WindowBuildError import WindowBuildError
from file_importing import import_edf_from_path
from alex_raw import alex_raw
from alex_raw_efficient import alex_raw_efficient
import pandas as pd
import datetime
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
import numpy as np
import re
import gc
import glob
from utils import config
vars_dict = config.get_vars_dict()

def record_file_info(file_path, EEG_times_arr, fis, save_dir_base, chunk, stitched_path="ERROR"):
    """
    Records file information and timestamps in the chunks file.
    
    Args:
        file_path (str): Path to the file being processed
        EEG_times_from_analyze_end_times (array): Array of timestamps
        fis (str): FIS identifier
    """
    excel_path = save_dir_base + r"\nicolet_raw_file_timestamps.xlsx"
    if not os.path.exists(excel_path):
        df = pd.DataFrame(columns=['file_name', 'fis', 'start_time', 'end_time'])
        df.to_excel(excel_path, index=False)
    df = pd.read_excel(excel_path)    
    start_time = EEG_times_arr[0]
    end_time = EEG_times_arr[1]
    end_stitched = os.path.split(file_path)[1]+"_stitched_cut.pkl"
    start_stitched = os.path.split(stitched_path)[0]
    stitched_path = os.path.join(start_stitched,end_stitched)

    new_row = {
        'file_name': file_path,
        'fis': fis,
        'start_time': start_time,
        'end_time': end_time,
        'chunk' : chunk,
        'stitched_path' : stitched_path
    }
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_excel(excel_path, index=False)
    print(f"Added {file_path} (FIS{fis}) to {excel_path}")

def initialize_error_files(save_dir_base, clear=True):
    """
    Helper function that initializes the Excel file error_log_OBM.xlsx and the text file error_log_OBM.txt.
    """
    # File names
    excel_file_name = save_dir_base + r"\nicolet_error_log.xlsx"
    text_file_name = save_dir_base + r"\nicolet_error_log.txt"
    today_date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    #Step 1: Initialize the Excel file
    # Create the files_info sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["file_path", "error_type"])  # Example header row
    # Save the workbook
    wb.save(excel_file_name)
    print(f"Excel file '{excel_file_name}' initialized with default values.")

    #Step 2: Initialize or clear the text file
    if clear or not os.path.exists(text_file_name):
        with open(text_file_name, "w") as log_file:
            log_file.write(f"{today_date_time}: error log initialized.\n")
        print(f"Text file '{text_file_name}' initialized with the current date and time.")
    else:
        print(f"Text file '{text_file_name}' already exists. No changes made.")
    
def record_error(error_string, save_dir_base):
    """
    Helper function that builds an error log for this fxn.

    Inputs:
    - error_string (str): the string we want added to the error log
    """
    # Define the log file path
    log_file_path = save_dir_base + r"\nicolet_error_log.txt"
    with open(log_file_path, "a") as log_file:
        # Write log initialization with today's date
        log_file.write(f"{error_string}\n\n")


def add_error_to_excel(error_type, file_path, save_dir_base):
    """
    Helper function that updates the information in error_log.xlsx any time there is an error.
    """
    file_name = save_dir_base + r"\nicolet_error_log.xlsx"
    wb = openpyxl.load_workbook(file_name)
    files_info_sheet = wb.active
    next_row = files_info_sheet.max_row + 1
    files_info_sheet.cell(row=next_row, column=1, value=file_path)  # Add file_path
    files_info_sheet.cell(row=next_row, column=2, value=error_type)  # Add error_type
    wb.save(file_name)
    print(f"Error of type '{error_type}' logged for file: {file_path}")
    
def clean_up_logs_and_excel(save_dir_base):
    # File paths
    error_log_txt = save_dir_base + r"\nicolet_error_log.txt"
    error_log_xlsx = save_dir_base + r"\nicolet_error_log.xlsx"
    all_times_excel = save_dir_base + r"\nicolet_raw_file_timestamps.xlsx"
    del_arr = [error_log_txt, error_log_xlsx, all_times_excel]
    for dir in del_arr:
        try:
            os.remove(dir)
        except FileNotFoundError:
            pass
    initialize_error_files(save_dir_base)

    #Clear directories
    files = glob.glob(os.path.join(save_dir_base,r'stitched_cut_efficient\*'))
    for f in files:
        os.remove(f)

def stitch_EDF_and_cut(e_file_directory, save_dir_base):
    """
    Core pipeline function that stitches .edf files and extracts time windows.
    
    This function processes all .edf files in a FIS directory, combining them into
    a single continuous EEG recording while preserving pause information and 
    extracting only specified analysis time windows.
    
    Parameters:
    -----------
    e_file_directory : str
        Full path to directory containing .edf files (e.g., 'path/to/FIS79post2/')
        Must contain at least one .edf file with valid annotations
    save_dir_base : str  
        Base directory for saving output files and logs
        
    Returns:
    --------
    None or 0
        Returns 0 on error (corrupted files, missing data, conversion failures)
        Returns None on successful completion
        
    Processing Steps:
    ----------------
    1. Sort .edf files chronologically using custom_sort()
    2. For each .edf file:
       - Extract pause duration from annotations (HH:MM:SS format)
       - Convert to alex_raw object
       - Add zero-padding for pause periods (NB: pauses between surgery end time and signal start are added much later in the processing pipeline)
       - Concatenate voltage data
    3. Build timestamp arrays (from zero, from surgery end)
    4. Extract datetime information from Excel metadata
    5. Cut data to specified time windows only
    6. Save as alex_raw_efficient object
    
    Error Conditions:
    ----------------
    - Corrupted .edf files: Logs error, returns 0
    - Missing timestamp data: Logs error, returns 0  
    - alex_raw conversion failure: Logs error, returns 0
    - Window building failure: Logs error, returns 0
    
    Side Effects:
    ------------
    - Updates Excel tracking files
    - Creates error logs
    - Performs garbage collection
    """
    #Good to know: EDF files are indeed split in order. In the annotations, the amount of time provided is the amount of time PRIOR to the beginning of the signal (for a given EDF file) that acts as padding/paused.

    
    gc.collect()
    
    print("Stitching and cutting files.")
    nicolet_channel_choice = vars_dict["nicolet_channels"]
    windows = vars_dict["windows"]
    recording_time_data_path = vars_dict["nicolet_recording_time_data_path"]
    surgery_excel = vars_dict["surgery_times_path"]

    voltages_arr = [] #This will be an array of arrays, each of which contains voltage data from the pause or  the edf file
    times_arr = [] #At the end, srate will be used to build this out starting from 0 sec.
    paths_arr = [] #This will be used to verify that files were stitched or concatenated in the correct order

    final_file_name = str(os.path.basename(os.path.normpath(e_file_directory))) #This DOES NOT include .edf; e.g., will be FIS101post2

    def custom_sort(file_list):
        """
        Sort .edf files by filename length then numeric suffix for chronological order.
        
        Nicolet exports create numbered .edf files that need proper chronological
        sorting (e.g., file.edf, file2.edf, file10.edf should sort as 1,2,10 not 1,10,2).
        
        Parameters:
        -----------
        file_list : list of str
            List of .edf filenames to sort
            
        Returns:
        --------
        list of str
            Sorted filenames in chronological order
            
        Sorting Logic:
        -------------
        1. Files without numeric suffix get priority (-1)
        2. Sort by filename length first
        3. Then by numeric suffix value
        """
        def sort_key(filename):
            # Extract the numeric suffix before `.edf`, defaulting to 0 if none exists
            match = re.search(r'(\d+)\.edf$', filename)
            num_suffix = int(match.group(1)) if match else -1  # -1 for non-numeric suffix to come first
            # Return a tuple with length and numeric suffix to sort by
            return (len(filename), num_suffix)

        # Sort the list with the custom sort key
        return sorted(file_list, key=sort_key)

    def find_first_fis_instance(e_file_str):
        """
        Finds the first instance of a string formatted as "FISXXX" in e_file_directory.
        
        Parameters:
        - e_file_directory (str): The input string to search.
        
        Returns:
        - str: The first matched "FISXXX" string, or None if no match is found.
        """
        # Define the regex pattern: "FIS" followed by digits
        pattern = r"(?i)FIS\d+"
        # Search for the first match
        match = re.search(pattern, e_file_str)
        # Return the matched string, or None if no match
        return match.group(0) if match else None

    sorted_files = custom_sort(os.listdir(e_file_directory))
    pause_indices = []
    signal_length = 0
    pause_length_excel = 0
    segment_i = 0

    for edf_file in sorted_files:
        if edf_file.endswith('.edf'):
            segment_i = segment_i + 1
            edf_path = os.path.join(e_file_directory, edf_file) #E.g., a6c6f55e-34aa-4280-8f73-bed63e619fb2FIS101post2\FIS101post22.edf
            #
            try:
                raw_EDF = import_edf_from_path(str(edf_path), downsample=vars_dict["downsample_rate"])
                #raw_EDF = import_edf_from_path(str(edf_path))
            except ValueError as e:
                error_str = f"Error importing file: {edf_path}. The specific error was {e}. Since this error was simply with importing the edf file, this was likely an issue with the file being corrupted. The file {edf_path} was NOT saved."
                add_error_to_excel("corrupt",e_file_directory,save_dir_base)
                record_error(error_str, save_dir_base)
                record_file_info(e_file_directory, [0,0], "ERROR", save_dir_base, "-")
                return 0 #We terminate these since we can't realistically stitch when one of the EDF files is corrupted
            annotations = raw_EDF.annotations
            annotations_descriptions = annotations.description
        else:
            continue #Since it's not an edf file (and therefore can't be processed here)
        pause_data = False #Turns true once we have data regarding a pause (i.e., will remain false when examining the first .edf in a .e folder, as this will have no pause/event information)
        pause_length_seconds = 0
        paths_arr.append(edf_path)
        
        descriptions_with_time_format = [description for description in annotations_descriptions if re.search(r'\b\d{1,2}:\d{2}:\d{2}\b', description)]
        # Throw an error if more than one description contains a colon
        if len(descriptions_with_time_format) > 1:
            raise ValueError(f"More than one description in annotations_descriptions contains a colon (':'). The annotations were {annotations}, and the descriptions for the annotations were {annotations_descriptions}.")

        for description in annotations_descriptions:
            # Check if the description contains a time format "XX:XX:XX" (HH:MM:SS)
            if re.search(r'\b\d{1,2}:\d{2}:\d{2}\b', description):
                # Extract the time part from the description using string manipulation
                time_str = ''.join(filter(lambda x: x.isdigit() or x == ":", description))
                try:
                    # Convert extracted time string to datetime.time() using strptime
                    pause_datetime = datetime.strptime(time_str, '%H:%M:%S').time()
                    # Convert the time (hours, minutes, seconds) to total seconds
                    pause_length_seconds = timedelta(hours=pause_datetime.hour, 
                                                    minutes=pause_datetime.minute, 
                                                    seconds=pause_datetime.second).total_seconds()

                    # Now you have the pause length in seconds
                    print(f"Pause found: {pause_datetime} equals {pause_length_seconds} seconds")
                    pause_data = True
                    break #Assumes one pause per file; error handling has been added to manage this

                except ValueError:
                    print(f"Invalid time format in description: {description}")
                    raise ValueError
            elif segment_i == 1:
                pause_length_seconds = 0
        
        #The pause length (in SECONDS) is now stored in pause_length_seconds
        voltage_add_arr = []
        try:
            alex_raw_EDF = alex_raw(raw_EDF, which_channel=nicolet_channel_choice, filePath=str(edf_path), call_from="postop_file_finder")
        except Exception as e:
            add_error_to_excel("conversion",e_file_directory, save_dir_base)
            record_error(f"Error converting the file {e_file_directory} to an alex_raw object. The specific error raised was: {e}. Therefore, the file was NOT saved.", save_dir_base)
            record_file_info(e_file_directory, [0,0], "ERROR", save_dir_base, "-")
            return 0
        alex_raw_EDF.OBM = False
        sfreq = alex_raw_EDF.sampling_rate #(in Hz)
        print(f"Sampling rate: {int(sfreq)} Hz.")
        n_data_points_to_add = int(sfreq * pause_length_seconds)
        edf_voltage_arr = np.round(alex_raw_EDF.EEG_data, decimals=2).astype(np.float32)
        voltage_add_arr = [0] * n_data_points_to_add

        if pause_data: #This adds pause length of fillvar in those cases where there was a pause
            voltages_arr.extend(voltage_add_arr)
            pause_indices.append(len(voltages_arr)-1)
            voltages_arr.extend(edf_voltage_arr)
            print(f"Pause added to pause_indices. We have {len(pause_indices)} pauses recorded.")
            print(f"Stitched recording length thus far: {len(voltages_arr)/sfreq} sec.\n---")
            signal_length += len(edf_voltage_arr)
            pause_length_excel += len(voltage_add_arr)
        else:
            voltages_arr.extend(edf_voltage_arr)
            print(f"Stitched recording length thus far: {len(voltages_arr)/sfreq} sec.\n---")
            signal_length += len(edf_voltage_arr)
    
    if segment_i == 0: #This means this directory had no .edf files, so we return 0 to break the function (which usually doesn't return anything)
        raise ValueError(f"The directory provided ({e_file_directory}) has no .edf files.")
    
    # Save post-processing data to Excel sheet
    wb = openpyxl.load_workbook(recording_time_data_path)
    sheet = wb["Sheet1"]
    # Check if necessary columns exist; if not, create them
    columns_to_add = ["stitched_time_hrs", "signal_length_hrs", "pause_length_hrs"]
    header = [cell.value for cell in sheet[1]]
    for col_name in columns_to_add:
        if col_name not in header:
            sheet.cell(row=1, column=len(header) + 1).value = col_name
            header.append(col_name)
    # Locate the row where the last part of "folder_path" equals the last part of e_file_directory
    row_of_interest = None
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        folder_path_value = row[header.index("folder_path")]
        # Extract the last part of folder_path and e_file_directory
        folder_name = folder_path_value.split('\\')[-1] if folder_path_value else ""
        e_file_directory_name = e_file_directory.split('\\')[-1]
        if folder_name == e_file_directory_name:
            row_of_interest = idx  # `idx` is now the actual row number in the sheet
            break

    # Extract recording time from "recording_time_total" and round to 1 decimal
    if row_of_interest is not None:
        # Add calculated values to the respective columns
        stitched_time_hrs = round(len(voltages_arr) / alex_raw_EDF.sampling_rate / 3600, 5)
        signal_length_hrs = round(signal_length / alex_raw_EDF.sampling_rate / 3600, 5)
        pause_length_hrs = round(pause_length_excel / alex_raw_EDF.sampling_rate / 3600, 5)
        # Insert values at row_of_interest in respective columns
        sheet.cell(row=row_of_interest, column=header.index("stitched_time_hrs") + 1, value=stitched_time_hrs)
        sheet.cell(row=row_of_interest, column=header.index("signal_length_hrs") + 1, value=signal_length_hrs)
        sheet.cell(row=row_of_interest, column=header.index("pause_length_hrs") + 1, value=pause_length_hrs)
    
    print("Building arrays of all times. This may take a minute...")
    times_arr = np.arange(len(voltages_arr)) / sfreq

    print("Copying the last alex_raw_EDF object...")
    alex_raw_stitched_and_cut = alex_raw_EDF

    print("Copying voltages to the new alex_raw object...")
    alex_raw_stitched_and_cut.EEG_data = voltages_arr
    del voltages_arr
    print("Copying times to the new alex_raw object...")
    alex_raw_stitched_and_cut.EEG_times_from_zero = times_arr
    del times_arr
    alex_raw_stitched_and_cut.pause_indices_stitched_and_cut = pause_indices
    #-------------------------------------------------------------------------
    #Add datetime information to our alex_raw_stitched object
    if not (windows == [["surgery"]] or windows == [["CPB"]]):
        file_path = recording_time_data_path
        num_fis = int(alex_raw_stitched_and_cut.num_fis)
        print(f"Matching up num_fis: {num_fis}")

        print("Reading Excel for times...")
        df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl', dtype={
        'recording_start_date': str,
        'recording_start_time': str,
        'date_cx': str,
        'analyze_end_time': str})
        print("Extracting times of interest")

        #Get row of interest to match
        df['folder_name'] = df['folder_path'].apply(lambda x: x.split('\\')[-1])
        e_file_directory_name = e_file_directory.split('\\')[-1]
        rowi = df.index[df['folder_name'] == e_file_directory_name].tolist()[0]

        rec_start_date = str(df.at[rowi, 'recording_start_date']).zfill(8)
        try:
            rec_start_date = datetime.strptime(rec_start_date, '%d%m%Y').date()
        except ValueError as e:
            error_str= f"For the file {e_file_directory}, we got the error: {e} when reading the datetime."
            #print(error_str)
            add_error_to_excel("time",e_file_directory,save_dir_base)
            record_error(error_str, save_dir_base)
            record_file_info(e_file_directory, [0, 0], alex_raw_stitched_and_cut.num_fis, save_dir_base, "-")
            return 0
        rec_start_time = str(df.at[rowi, 'recording_start_time']).zfill(6)
        rec_start_time = datetime.strptime(rec_start_time, '%H%M%S').time()
        rec_start_date_time = datetime.combine(rec_start_date, rec_start_time)
        alex_raw_stitched_and_cut.rec_start_date_time = rec_start_date_time
        
        #Get date of surgery
        clindata_file_path = recording_time_data_path
        # Read the Excel sheet into a DataFrame
        df = pd.read_excel(clindata_file_path)
        df['fis_num'] = df['fis_num'].astype(int)
        # Find the row where 'fis_num' matches 'num_fis'
        row = df[df['fis_num'] == num_fis]
        if not row.empty:
            # Extract the value in the column 'Fecha Cx'
            fecha_cx_str = row.iloc[0]['date_cx']
            # Convert the string to a datetime object
            fecha_cx_str = fecha_cx_str.strftime('%d/%m/%Y')  # Convert Timestamp to string
            cx_date = datetime.strptime(fecha_cx_str, '%d/%m/%Y')
        else:
            raise ValueError(f"No row for FIS{alex_raw_stitched_and_cut.fis} found in the sheet clinical_Data_Only.xlsx")

        if isinstance(df.at[rowi, 'analyze_end_time'], (int, float)) and np.isnan(df.at[rowi, 'analyze_end_time']):
            add_error_to_excel("time",e_file_directory,save_dir_base)
            record_error(f"There was no surgery time recorded in all_nicolet_folders_cleaning.xlsx for the file at {e_file_directory}.", save_dir_base)
            record_file_info(e_file_directory, [0,0], "ERROR", save_dir_base, "-")
            return 0
        analyze_time = str(df.at[rowi, 'analyze_end_time'])
        analyze_time = analyze_time.split('.')[0]
        analyze_time = datetime.strptime(analyze_time, '%H%M').time()
        analyze_date_time = datetime.combine(cx_date, analyze_time)
        alex_raw_stitched_and_cut.rec_start_date_time = rec_start_date_time
        alex_raw_stitched_and_cut.cx_date = cx_date
        alex_raw_stitched_and_cut.analyze_end_date_time = analyze_date_time
        alex_raw_stitched_and_cut.window_type = "postop"
    else:
        #Step 1: Access/extract information about the start of the recording
        file_path = recording_time_data_path
        num_fis = int(alex_raw_stitched_and_cut.num_fis)
        print(f"Matching up num_fis: {num_fis}")
        print("Reading Excel for times...")
        df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl', dtype={
        'recording_start_date': str,
        'recording_start_time': str,
        'date_cx': str})
        print("Extracting times of interest")
        df['folder_name'] = df['folder_path'].apply(lambda x: x.split('\\')[-1])
        e_file_directory_name = e_file_directory.split('\\')[-1]
        rowi = df.index[df['folder_name'] == e_file_directory_name].tolist()[0]
        rec_start_date = str(df.at[rowi, 'recording_start_date']).zfill(8)
        try:
            rec_start_date = datetime.strptime(rec_start_date, '%d%m%Y').date()
        except ValueError as e:
            error_str= f"For the file {e_file_directory}, we got the error: {e} when reading the datetime."
            #print(error_str)
            add_error_to_excel("time_surg_CPB",e_file_directory,save_dir_base)
            record_error(error_str, save_dir_base)
            record_file_info(e_file_directory, [0,0], "ERROR", save_dir_base, "-")
            return 0
        rec_start_time = str(df.at[rowi, 'recording_start_time']).zfill(6)
        rec_start_time = datetime.strptime(rec_start_time, '%H%M%S').time()
        rec_start_date_time = datetime.combine(rec_start_date, rec_start_time)
        alex_raw_stitched_and_cut.rec_start_date_time = rec_start_date_time

        #Step 2: Get the date of the surgery
        #Get date of surgery
        surgery_info_file_path = surgery_excel
        # Read the Excel sheet into a DataFrame
        df_surg = pd.read_excel(surgery_info_file_path)
        df_surg['fis_num'] = df_surg['fis_num'].astype(int)
        # Find the row where 'fis_num' matches 'num_fis'
        row = df_surg[df_surg['fis_num'] == num_fis]
        if not row.empty:
            # Extract the value in the column 'Fecha Cx'
            fecha_cx_str = row.iloc[0]['date_surgery']
            # Convert the string to a datetime object
            fecha_cx_str = fecha_cx_str.strftime('%d/%m/%Y')  # Convert Timestamp to string
            cx_date = datetime.strptime(fecha_cx_str, '%d/%m/%Y')
        else:
            raise ValueError(f"No row for FIS{alex_raw_stitched_and_cut.fis} found in the sheet surgery_time_data.xlsx")
        
        time_end_val = row.iloc[0]['time_end_cx']
        time_start_val = row.iloc[0]['time_start_cx']

        if isinstance(time_end_val, (int, float)) and np.isnan(time_end_val) or isinstance(time_start_val, (int, float)) and np.isnan(time_start_val):
            add_error_to_excel("time",e_file_directory,save_dir_base)
            record_error(f"There was no surgery start or end time recorded in surgery_time_data.xlsx for the file at {e_file_directory}.", save_dir_base)
            record_file_info(e_file_directory, [0,0], "ERROR", save_dir_base, "-")
            return 0
        
        #Step 3: set times
        if windows == [["surgery"]]:
            analyze_time = str(time_end_val)
            analyze_time = analyze_time.split('.')[0]
            analyze_time = datetime.strptime(analyze_time, '%H%M').time()
            analyze_date_time = datetime.combine(cx_date, analyze_time)
            analyze_start_time = str(time_start_val)
            analyze_start_time = analyze_start_time.split('.')[0]
            analyze_start_time = datetime.strptime(analyze_start_time, '%H%M').time()
            analyze_start_date_time = datetime.combine(cx_date, analyze_start_time)
            alex_raw_stitched_and_cut.rec_start_date_time = rec_start_date_time
            alex_raw_stitched_and_cut.cx_date = cx_date
            alex_raw_stitched_and_cut.analyze_end_date_time = analyze_date_time
            alex_raw_stitched_and_cut.analyze_start_date_time = analyze_start_date_time
            alex_raw_stitched_and_cut.window_type = "postop"
        elif windows == [["CPB"]]:
            time_end_val = row.iloc[0]['time_end_ecc']
            time_start_val = row.iloc[0]['time_start_ecc']
            if (isinstance(time_end_val, (int, float)) and np.isnan(time_end_val)) or (isinstance(time_start_val, (int, float)) and np.isnan(time_start_val)) or time_start_val == "-" or time_end_val == "-":
                add_error_to_excel("time_CPB",e_file_directory,save_dir_base)
                record_error(f"There was no CPB start or end time recorded in surgery_time_data.xlsx for the file at {e_file_directory}.", save_dir_base)
                record_file_info(e_file_directory, [0,0], "ERROR", save_dir_base, "-")
                return 0
            analyze_time = str(time_end_val)
            analyze_time = analyze_time.split('.')[0]
            analyze_time = datetime.strptime(analyze_time, '%H%M').time()
            analyze_date_time = datetime.combine(cx_date, analyze_time)
            analyze_start_time = str(time_start_val)
            analyze_start_time = analyze_start_time.split('.')[0]
            analyze_start_time = datetime.strptime(analyze_start_time, '%H%M').time()
            analyze_start_date_time = datetime.combine(cx_date, analyze_start_time)
            alex_raw_stitched_and_cut.rec_start_date_time = rec_start_date_time
            alex_raw_stitched_and_cut.cx_date = cx_date
            alex_raw_stitched_and_cut.analyze_end_date_time = analyze_date_time
            alex_raw_stitched_and_cut.analyze_start_date_time = analyze_start_date_time
            alex_raw_stitched_and_cut.window_type = "CPB"





    # Build timestamps relative to surgery end time for analysis
    # This allows analysis relative to specific surgical events
    print("Building array of times relative to end of surgery...")
    rec_start_date_time = alex_raw_stitched_and_cut.rec_start_date_time
    analyze_end_date_time = alex_raw_stitched_and_cut.analyze_end_date_time
    sampling_rate = alex_raw_stitched_and_cut.sampling_rate
    eeg_times_length = len(alex_raw_stitched_and_cut.EEG_times_from_zero)
    # Calculate offset: negative if recording started before surgery end
    initial_diff_seconds = (rec_start_date_time - analyze_end_date_time).total_seconds()

    # Generate timestamp array: each sample gets time relative to surgery end
    time_deltas = np.arange(eeg_times_length) / sampling_rate
    seconds_from_end_date_time = initial_diff_seconds + time_deltas

    alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time = np.array(seconds_from_end_date_time) #NB: this is stored as SECONDS

    current_path = alex_raw_stitched_and_cut.file_path_current
    file_name_no_pkl = os.path.splitext(os.path.basename(current_path))[0]
    new_file_name = file_name_no_pkl+"_stitched_cut"+'.pkl'
    new_path = os.path.join(save_dir_base + r"\stitched_cut_efficient", new_file_name)
    original_times = deepcopy([alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[0], alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[-1]])
    
    #Debugging
    print(f"rec_start_date_time: {rec_start_date_time}")
    print(f"analyze_end_date_time: {analyze_end_date_time}")
    print(f"initial_diff_seconds: {initial_diff_seconds}")
    print(f"First few times: {seconds_from_end_date_time[:5]}")
    
    del time_deltas
    del seconds_from_end_date_time 

    pause_times_postop = []
    print(f"We have {len(pause_indices)} pauses recorded.")
    for pause_index in pause_indices:
        pause_times_postop.append(round(alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[pause_index] / 3600, 2))
    print (f"The pauses occurred at the following number of hours postop: {pause_times_postop}\n\n")
    alex_raw_stitched_and_cut.pause_times_postop_hrs = pause_times_postop

    #-------------------------------------------------------------------------
    #Re-build our alex_raw.EEG_data, alex_raw.EEG_times_from_zero, and alex_raw.EEG_times_from_analyze_end_time arrays to contain only data in the windows of interest. Additionally, store the windows and a stitched_and_cut boolean.
    print("Cutting down data...")
    if not (windows == [["surgery"]] or windows == [["CPB"]]):
        try:
            window_indices, windows = alex_raw_stitched_and_cut.get_window_indices(windows) #window_indices are the pairs of indices corresponding to each window in windows (e.g., if window_indices is [a,b], then alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[a] corresponds to the seconds values at 20 hr (if windows is [[20,24]]))
        except WindowBuildError as e:
            error_str= f"For the file {e_file_directory}, we got the error: {e} when trying to build windows."
            #print(error_str)
            add_error_to_excel("window",e_file_directory,save_dir_base)
            record_error(error_str, save_dir_base)
            record_file_info(e_file_directory, [alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[0], alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[1]], "ERROR", save_dir_base, "-")
            return 0
    else:
        windows = [[(alex_raw_stitched_and_cut.analyze_start_date_time - alex_raw_stitched_and_cut.analyze_end_date_time).total_seconds() / 3600, 0]]
        print(f"Our surgery times are {windows[0][0]}-{windows[0][1]} hrs.")
        try:
            window_indices, windows = alex_raw_stitched_and_cut.get_window_indices(windows) #window_indices are the pairs of indices corresponding to each window in windows (e.g., if window_indices is [a,b], then alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[a] corresponds to the seconds values at 20 hr (if windows is [[20,24]]))
        except WindowBuildError as e:
            error_str= f"For the file {e_file_directory}, we got the error: {e} when trying to build windows."
            #print(error_str)
            add_error_to_excel("window",e_file_directory,save_dir_base)
            record_error(error_str, save_dir_base)
            record_file_info(e_file_directory, [alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[0], alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[1]], "ERROR", save_dir_base, "-")
            return 0

    window_lengths_secs = []
    my_windows = []
    data_slices = []
    times_from_zero_slices = []
    times_analyze_slices = []
    i = 0
    # Extract data slices for each valid time window
    # window_indices contains [start_idx, end_idx] pairs for each window
    for window in window_indices: #This is only for windows with data. NB: window here is a window index in alex_raw.EEG_data (or EEG_times_from_zero, etc)
        index1 = window[0]
        index2 = window[1]
        # Handle edge case where window extends to end of data
        if (index2+1)<len(alex_raw_stitched_and_cut.EEG_data):
            # Normal case: extract full window including end point
            data_slices.append(alex_raw_stitched_and_cut.EEG_data[index1:index2+1])
            times_from_zero_slices.append(alex_raw_stitched_and_cut.EEG_times_from_zero[index1:index2+1])
            times_analyze_slices.append(alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[index1:index2+1])
        else:
            # Edge case: window extends to data end, exclude +1 to avoid index error
            data_slices.append(alex_raw_stitched_and_cut.EEG_data[index1:index2])
            times_from_zero_slices.append(alex_raw_stitched_and_cut.EEG_times_from_zero[index1:index2])
            times_analyze_slices.append(alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time[index1:index2])
        my_windows.append(windows[i])
        window_length = times_from_zero_slices[i][-1] - times_from_zero_slices[i][0] + 1/alex_raw_stitched_and_cut.sampling_rate
        window_lengths_secs.append(window_length) #Each window 
        i = i + 1
    
    alex_raw_stitched_and_cut.EEG_data = np.concatenate(data_slices)
    del data_slices
    alex_raw_stitched_and_cut.EEG_times_from_zero = np.concatenate(times_from_zero_slices)
    del times_from_zero_slices
    alex_raw_stitched_and_cut.EEG_times_from_analyze_end_time = np.concatenate(times_analyze_slices)
    del times_analyze_slices

    alex_raw_stitched_and_cut.windows = my_windows #This way the object only retains the window it contains
    alex_raw_stitched_and_cut.stitched_and_cut = True
    alex_raw_stitched_and_cut.window_indices, alex_raw_stitched_and_cut.windows = alex_raw_stitched_and_cut.get_window_indices(alex_raw_stitched_and_cut.windows) #Re-assigning after we've updated the data
    alex_raw_stitched_and_cut.window_lengths_secs = window_lengths_secs
        
    stitched_file_name=final_file_name #E.g., FIS101post2
    alex_raw_stitched_and_cut.saved_once = False
    alex_raw_stitched_and_cut.file_path_current = os.path.join(r"E:\EEG_Pipeline_Continuous\1_stitched_cut\stitched_cut_full_files",stitched_file_name) #Because our file name is changing, we have to manually update self.file_path_current
    alex_raw_stitched_and_cut.add_change_log(f"Data was stitched together from the files {paths_arr} in the order shown. Pauses were retained, and their duration was marked with {0} in the object. Furthermore, the data was parsed to only include times from the windows: {windows}; of these, the windows that were found in this file were {alex_raw_stitched_and_cut.windows}, which was indeed in line with the Excel file.")
    print("Stitched file built. Now saving...")
    #alex_raw_stitched_and_cut.save_as(r"E:\EEG_Pipeline_Continuous\1_stitched_cut\stitched_cut_full_files","_stitched_cut")
    print("Creating alex_raw_efficient object...")
    alex_raw_small = alex_raw_efficient(alex_raw_stitched_and_cut)
    alex_raw_small.save_as_easy(save_dir_base + r"\stitched_cut_efficient","_stitched_cut")
    record_file_info(e_file_directory, original_times, alex_raw_stitched_and_cut.num_fis, save_dir_base, alex_raw_stitched_and_cut.windows[0], stitched_path = new_path)
    print("Moving on to next file...")
    
    
    return None #Done so local variables clear

def is_nicolet(fis_folder, base_path=Path("E:/gimeno_Datos_aEEG/all_EDF_Files")):
    """
    This function takes in a FIS folder (e.g., FIS102b) and determines if this folder reflects a Nicolet file structure.
    A Nicolet structure contains at least one subdirectory with the FIS of interest.

    Inputs:
    - fis_folder (str): the name of the FIS folder of interest (e.g., "FIS102b")
    - base_path (Path or str): the base directory where the FIS folder is located. Default is "E:/gimeno_Datos_aEEG/all_EDF_Files".

    Outputs:
    - is_nicolet (boolean): True if the folder reflects Nicolet structure, False otherwise.
    """
    
    # Construct the full path by joining the base path and the fis_folder
    fis_path = Path(base_path) / fis_folder

    # Check if the FIS folder exists
    if not fis_path.exists():
        print(f"Directory {fis_path} does not exist.")
        return False
    
    # List subdirectories within the FIS folder
    subdirs = [subdir for subdir in fis_path.iterdir() if subdir.is_dir()]
    
    # If there is at least one subdirectory, it is considered a Nicolet structure
    is_nicolet = len(subdirs) > 0
    
    return is_nicolet

def iterate_stitch_n(save_dir_base):
    r""""
    Function to iterate over EDF files. It decides whether or not to process a file based on whether or not the recording_time column in the Excel file E:\EEG_Pipeline_Continuous\1_stitched_cut\time_data\all_nicolet_clean_MATLAB.xlsx has an entry. (I.e., if it does not have an entry, we process it).

    Inputs:
    - efficient (bool): if True, will use the directories in the file Nicolet_folders_in_range_summary.xlsx at C:\Users\user\OneDrive\Vandy\Research\Barcelona\EEG_Project\Code_EEG_Analysis\Sample_Characteristics to guide which are processed. Otherwise, will simply iterate through all folders at .
    - cut (bool): if True, will retain only data in the winows of interest (default); otherwise, will retian all data. Obviously, maintaining data in the windows of interest is more efficient and leads to faster loading/saving.
    """
    os.makedirs(rf"{save_dir_base}\stitched_cut_efficient", exist_ok = True)
    os.makedirs(rf"{save_dir_base}\stitched_cut_final", exist_ok = True)
    clean_up_logs_and_excel(save_dir_base)
    for subdirectory in os.listdir(vars_dict["nicolet_folder"]):
        subdirectory_path = os.path.join(vars_dict["nicolet_folder"], subdirectory)
        stitch_EDF_and_cut(subdirectory_path, save_dir_base)


